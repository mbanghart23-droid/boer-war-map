"""
Comprehensive data quality audit of movements.csv + events.geojson.
Checks every category of issue, not just chronological gaps.
"""
import csv, json, re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent.parent

rows = list(csv.DictReader(open(HERE/"data"/"movements.csv", encoding="utf-8")))
gj   = json.load(open(HERE/"docs"/"data"/"events.geojson", encoding="utf-8"))

row_by_id = {r["id"]: r for r in rows}
feats = {f["properties"]["id"]: f for f in gj["features"] if f["properties"]["kind"]=="event"}

def parse_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], "%Y-%m-%d")
    except: return None

WAR_START = datetime(1899, 10, 11)
WAR_END   = datetime(1902, 5, 31)

issues = []  # list of dicts

def add(category, severity, row_id, field, value, description):
    issues.append({
        "category": category,
        "severity": severity,   # CRITICAL / HIGH / MEDIUM / LOW
        "id": row_id,
        "field": field,
        "value": str(value)[:80],
        "description": description,
        "force": row_by_id.get(row_id,{}).get("force","")[:50],
    })

# ── 1. ID coverage ────────────────────────────────────────────────────────────
all_ids = [int(r["id"]) for r in rows if r["id"].isdigit()]
id_max = max(all_ids)
id_set = set(all_ids)
for i in range(1, id_max+1):
    if i not in id_set:
        add("ID Gaps", "LOW", str(i), "id", i, f"Row ID {i} is missing from CSV (deleted or never used)")

# ── 2. Rows NOT in GeoJSON (no A dict entry) ─────────────────────────────────
geojson_ids = set(feats.keys())
for r in rows:
    if r["id"] not in geojson_ids:
        add("Not on Map", "HIGH", r["id"], "A dict", "", f"Row has no map placement (not in A dict): {r['force'][:50]} / {r['action_place'] or r['from_place']}")

# ── 3. Date issues ────────────────────────────────────────────────────────────
for r in rows:
    rid = r["id"]
    ds = parse_date(r.get("date_start",""))
    de = parse_date(r.get("date_end",""))

    if not ds:
        add("Date", "HIGH", rid, "date_start", r.get("date_start",""), "Missing or unparseable date_start")

    if ds and de and de < ds:
        add("Date", "CRITICAL", rid, "date_end", r.get("date_end",""), f"date_end {r['date_end']} is BEFORE date_start {r['date_start']}")

    if ds and (ds < WAR_START or ds > WAR_END):
        add("Date", "MEDIUM", rid, "date_start", r.get("date_start",""), f"Date {r['date_start']} is outside war period (1899-10-11 to 1902-05-31)")

    if ds and de:
        delta = (de - ds).days
        if delta > 365:
            add("Date", "MEDIUM", rid, "date_range", f"{r['date_start']}–{r['date_end']}", f"Date range is {delta} days — likely a deployment placeholder, not a single event")

# ── 4. Same-day place conflict (unit in two places same day) ──────────────────
# Group by unit + date
unit_date_place = defaultdict(list)
for r in rows:
    if not r.get("date_start"): continue
    key = (r.get("force",""), r["date_start"][:10])
    place = r.get("action_place") or r.get("from_place") or ""
    unit_date_place[key].append((r["id"], place))

for (force, date), entries in unit_date_place.items():
    places = set(p for _, p in entries if p)
    if len(places) > 1:
        ids = [i for i,_ in entries]
        add("Logic", "MEDIUM", ids[0], "place conflict",
            f"{date}: {', '.join(places)}",
            f"'{force}' appears in multiple places on {date}: {', '.join(places)} (IDs {', '.join(ids)})")

# ── 5. Missing descriptions / thin descriptions ───────────────────────────────
for r in rows:
    desc = r.get("description","").strip()
    if not desc:
        add("Description", "HIGH", r["id"], "description", "", "No description at all")
    elif len(desc) < 40:
        add("Description", "MEDIUM", r["id"], "description", desc, f"Very short description ({len(desc)} chars)")
    elif len(desc) < 80:
        add("Description", "LOW", r["id"], "description", desc[:60], f"Short description ({len(desc)} chars) — may need enrichment")

# ── 6. Missing source ─────────────────────────────────────────────────────────
for r in rows:
    src = r.get("source","").strip()
    if not src:
        add("Source", "MEDIUM", r["id"], "source", "", "No source cited")
    elif src.lower() in ("tbd","unknown","?","n/a","to be determined"):
        add("Source", "MEDIUM", r["id"], "source", src, "Placeholder source — needs real citation")

# ── 7. Province-level / vague places remaining ───────────────────────────────
VAGUE_PLACES = {"Cape Colony","Orange Free State","Orange River Colony","Natal","Transvaal",
                "South Africa","Eastern Cape","Western Cape","Northern Cape","North-west Cape",
                "Cape Midlands (intended)","Cape Midlands","unknown","Unknown","TBD"}
for r in rows:
    for field in ("action_place","from_place","to_place"):
        val = r.get(field,"").strip()
        if val in VAGUE_PLACES:
            add("Place", "HIGH", r["id"], field, val, f"Province/region-level place in {field} — needs specific town")
        elif re.search(r"\bdistrict\b|\bprovince\b|\bregion\b", val, re.I):
            add("Place", "MEDIUM", r["id"], field, val, f"Vague geographic term in {field}: '{val}'")

# ── 8. Commander format (should be "Lastname, Firstname" for British) ─────────
CMD_PATTERN = re.compile(r"^[A-Z][a-z]+,\s")  # "Smith, John" pattern
SKIP_CMD = {"","Lord Kitchener","Lord Roberts","Lord Methuen","Lord Dundonald",
            "Gen Joubert","Gen Cronjé","Gen Botha","Gen De Wet","Gen De la Rey",
            "unknown","Unknown","—"}
for r in rows:
    cmd = r.get("commander","").strip()
    if not cmd or cmd in SKIP_CMD: continue
    if r.get("side","") == "British":
        # Should match "Lastname, Firstname" or title forms
        if not CMD_PATTERN.match(cmd) and not cmd.startswith("Lord ") and not cmd.startswith("Sir "):
            add("Commander", "LOW", r["id"], "commander", cmd, f"British commander may not be in 'Lastname, Firstname' format: '{cmd}'")

# ── 9. Missing event_type ─────────────────────────────────────────────────────
VALID_TYPES = {"engagement","battle","raid","advance","retreat","movement","siege","capture",
               "surrender","defeat","pursuit","skirmish","garrison","execution","disembark",
               "rail_move","rail-move","redeployment","occupation","command","drive","deployment"}
for r in rows:
    et = r.get("event_type","").strip()
    if not et:
        add("Event Type", "HIGH", r["id"], "event_type", "", "No event_type")
    elif et not in VALID_TYPES:
        add("Event Type", "MEDIUM", r["id"], "event_type", et, f"Non-standard event_type: '{et}'")

# ── 10. Low geocoding confidence ──────────────────────────────────────────────
for fid, f in feats.items():
    p = f["properties"]
    if p.get("geo_confidence") == "low":
        add("Geocoding", "MEDIUM", fid, "geo_confidence", p.get("place",""),
            f"Low-confidence coordinates for '{p.get('place','')}' — marker may be significantly misplaced")

# ── 11. Missing side field ────────────────────────────────────────────────────
for r in rows:
    if r.get("side","").strip() not in ("British","Boer"):
        add("Side", "HIGH", r["id"], "side", r.get("side",""), f"Invalid side value: '{r.get('side','')}'")

# ── 12. Duplicate descriptions (copy-paste placeholders) ─────────────────────
desc_count = defaultdict(list)
for r in rows:
    d = r.get("description","").strip()
    if len(d) > 20:
        desc_count[d].append(r["id"])
for desc, ids in desc_count.items():
    if len(ids) > 1:
        add("Duplicate", "MEDIUM", ids[0], "description", desc[:60],
            f"Same description on {len(ids)} rows (IDs: {', '.join(ids)}) — possible copy-paste")

# ── 13. Rows where from_place == to_place (trivial movement) ─────────────────
for r in rows:
    fp = r.get("from_place","").strip()
    tp = r.get("to_place","").strip()
    if fp and tp and fp == tp and r.get("event_type") in ("advance","movement","raid","retreat"):
        add("Logic", "LOW", r["id"], "from/to", fp, f"from_place == to_place ('{fp}') on a movement row")

# ── 14. Movement rows without from/to (no arrow possible) ────────────────────
MOVEMENT_TYPES = {"advance","retreat","movement","raid","pursuit","drive","redeployment","rail_move","rail-move"}
for r in rows:
    if r.get("event_type") in MOVEMENT_TYPES:
        fp = r.get("from_place","").strip()
        tp = r.get("to_place","").strip()
        if not fp and not tp:
            add("Missing Arrow", "MEDIUM", r["id"], "from/to places", "",
                f"Movement-type row ('{r.get('event_type')}') has no from_place or to_place — can't draw arrow")

# ── 15. Confidence field check ────────────────────────────────────────────────
VALID_CONF = {"high","medium","low"}
for r in rows:
    c = r.get("confidence","").strip()
    if c not in VALID_CONF:
        add("Confidence", "LOW", r["id"], "confidence", c, f"Non-standard confidence value: '{c}'")

# ── 16. Units field missing on British rows ───────────────────────────────────
for r in rows:
    if r.get("side") == "British" and not r.get("units","").strip():
        add("Units", "LOW", r["id"], "units", "", "British row has no units field — won't appear in regiment filter correctly")

# ── 17. Events after peace of Vereeniging (31 May 1902) ──────────────────────
for r in rows:
    ds = parse_date(r.get("date_start",""))
    if ds and ds > WAR_END:
        add("Date", "HIGH", r["id"], "date_start", r.get("date_start",""),
            f"Event dated after Peace of Vereeniging (31 May 1902): {r.get('date_start','')}")

# ── summary ───────────────────────────────────────────────────────────────────
by_cat   = defaultdict(list)
by_sev   = defaultdict(int)
for iss in issues:
    by_cat[iss["category"]].append(iss)
    by_sev[iss["severity"]] += 1

print(f"\nTotal issues: {len(issues)}")
for sev in ["CRITICAL","HIGH","MEDIUM","LOW"]:
    print(f"  {sev}: {by_sev[sev]}")
print("\nBy category:")
for cat, lst in sorted(by_cat.items(), key=lambda x:-len(x[1])):
    print(f"  {cat}: {len(lst)}")

# ── write Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
SEV_COLOR = {"CRITICAL":"FF0000","HIGH":"FF6600","MEDIUM":"FFAA00","LOW":"CCCCCC"}

# ── Main issues sheet ─────────────────────────────────────────────────────────
ws = wb.active; ws.title = "All Issues"
HDR = ["Severity","Category","ID","Force","Field","Current Value","Issue Description"]
ws.append(HDR)
hf = PatternFill("solid",fgColor="1a1a2e"); hfn = Font(bold=True,color="FFFFFF",size=10)
for c in range(1,8): ws.cell(1,c).fill=hf; ws.cell(1,c).font=hfn

for iss in sorted(issues, key=lambda x: ["CRITICAL","HIGH","MEDIUM","LOW"].index(x["severity"])):
    ws.append([iss["severity"],iss["category"],iss["id"],iss["force"],iss["field"],iss["value"],iss["description"]])
    r = ws.max_row
    col = SEV_COLOR.get(iss["severity"],"CCCCCC")
    ws.cell(r,1).fill = PatternFill("solid",fgColor=col)
    ws.cell(r,1).font = Font(bold=True,color="FFFFFF" if iss["severity"] in ("CRITICAL","HIGH") else "333333")
    ws.cell(r,7).alignment = Alignment(wrap_text=True)

ws.column_dimensions["A"].width=10; ws.column_dimensions["B"].width=18
ws.column_dimensions["C"].width=6;  ws.column_dimensions["D"].width=35
ws.column_dimensions["E"].width=16; ws.column_dimensions["F"].width=30
ws.column_dimensions["G"].width=65
ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions

# ── Per-category summary sheet ────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.append(["Category","CRITICAL","HIGH","MEDIUM","LOW","TOTAL"])
hf2 = PatternFill("solid",fgColor="2c3e50"); hfn2 = Font(bold=True,color="FFFFFF",size=11)
for c in range(1,7): ws2.cell(1,c).fill=hf2; ws2.cell(1,c).font=hfn2
for cat, lst in sorted(by_cat.items()):
    counts = defaultdict(int)
    for i in lst: counts[i["severity"]]+=1
    ws2.append([cat,counts["CRITICAL"],counts["HIGH"],counts["MEDIUM"],counts["LOW"],len(lst)])

ws2.column_dimensions["A"].width=22
for col in "BCDEF": ws2.column_dimensions[col].width=10

# ── Not-on-map rows (most actionable) ────────────────────────────────────────
ws3 = wb.create_sheet("Not On Map")
ws3.append(["ID","Side","Force","Date Start","Event Type","Action Place","From Place","To Place","Description","Fix needed"])
hf3 = PatternFill("solid",fgColor="3d0000"); hfn3 = Font(bold=True,color="FFA0A0",size=10)
for c in range(1,11): ws3.cell(1,c).fill=hf3; ws3.cell(1,c).font=hfn3
not_on_map = [i for i in issues if i["category"]=="Not on Map"]
mapped_ids = set(geojson_ids)
for iss in not_on_map:
    r = row_by_id.get(iss["id"],{})
    ws3.append([r.get("id"),r.get("side"),r.get("force"),r.get("date_start"),r.get("event_type"),
                r.get("action_place"),r.get("from_place"),r.get("to_place"),
                r.get("description","")[:100],
                "Add to A dict in build_map.py with coords"])
for col in ["A","B","C","D","E","F","G","H","I","J"]:
    ws3.column_dimensions[col].width = [6,8,35,12,14,22,22,22,60,30]["ABCDEFGHIJ".index(col)]

# ── Thin descriptions (quick wins) ───────────────────────────────────────────
ws4 = wb.create_sheet("Thin Descriptions")
ws4.append(["ID","Side","Force","Date","Action Place","Event Type","Current Description","Chars","Fix"])
hf4 = PatternFill("solid",fgColor="1a3a00"); hfn4 = Font(bold=True,color="90FF90",size=10)
for c in range(1,10): ws4.cell(1,c).fill=hf4; ws4.cell(1,c).font=hfn4
thin = [i for i in issues if i["category"]=="Description" and i["severity"] in ("HIGH","MEDIUM")]
for iss in thin:
    r = row_by_id.get(iss["id"],{})
    desc = r.get("description","")
    ws4.append([r.get("id"),r.get("side"),r.get("force"),r.get("date_start"),
                r.get("action_place"),r.get("event_type"),desc,len(desc),"Needs enrichment"])
ws4.column_dimensions["A"].width=6; ws4.column_dimensions["B"].width=8
ws4.column_dimensions["C"].width=35; ws4.column_dimensions["D"].width=12
ws4.column_dimensions["E"].width=22; ws4.column_dimensions["F"].width=14
ws4.column_dimensions["G"].width=80; ws4.column_dimensions["H"].width=6
ws4.column_dimensions["I"].width=18

out = HERE/"tools"/"boer_war_full_audit.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
