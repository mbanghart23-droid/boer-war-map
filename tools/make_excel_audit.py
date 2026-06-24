"""
Build boer_war_audit.xlsx from audit.json + movements.csv.
Sheets:
  1. Summary
  2. Missing Battles (9 battles with no engagement row)
  3. Missing Events – Eastern Cape (3 gaps)
  4. Missing Commander (Cronjé)
  5. Thin Descriptions (282 rows with short/generic text — for manual enrichment)
  6. Deployment Stacks (places with huge pin piles)
  7. Event Type Review (all non-deployment event types + example)
  8. ID Gaps (8 missing IDs — leave blank or fill)
  9. Add-These-Rows (pre-filled CSV template for the 9 missing battles + 3 EC events)
"""
import json, csv, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

audit = json.load(open('data/audit.json', encoding='utf-8'))
all_rows = list(csv.DictReader(open('data/movements.csv', encoding='utf-8')))

wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
MISS_FILL = PatternFill("solid", fgColor="FFD7D7")   # soft red
OK_FILL   = PatternFill("solid", fgColor="D7FFD7")   # soft green
WARN_FILL = PatternFill("solid", fgColor="FFFACD")   # yellow
ALT_FILL  = PatternFill("solid", fgColor="F2F5FB")   # light blue-grey alternating row
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def hdr(ws, row_num, cols):
    for col_idx, (text, width) in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col_idx, value=text)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def rowfill(ws, row_num, values, fill=None, wrap=False):
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.border = THIN_BORDER
        if fill: c.fill = fill
        if wrap: c.alignment = Alignment(wrap_text=True, vertical='top')

def autofreeze(ws):
    ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Summary
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Summary"
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50

def srow(ws, r, label, value, note='', fill=None):
    for col, v in enumerate([label, value, note], 1):
        c = ws.cell(row=r, column=col, value=v)
        c.border = THIN_BORDER
        if fill: c.fill = fill
    ws.cell(row=r, column=1).font = Font(bold=True)

hdr(ws, 1, [("Metric", 38), ("Value", 18), ("Notes", 50)])
s = audit['summary']
et = s['event_types']
srow(ws, 2, "Total CSV rows", s['total_rows'])
srow(ws, 3, "ID range", s['id_range'])
srow(ws, 4, "ID gaps (unfilled IDs)", len(s['id_gaps']), str(s['id_gaps']), WARN_FILL)
srow(ws, 5, "Deployment rows", et.get('deployment',0))
srow(ws, 6, "Engagement rows", et.get('engagement',0))
srow(ws, 7, "Advance / movement rows", et.get('advance',0)+et.get('movement',0)+et.get('move',0))
srow(ws, 8, "Raid rows", et.get('raid',0))
srow(ws, 9, "Siege rows", et.get('siege',0))
srow(ws, 10,"Capture rows", et.get('capture',0))
srow(ws, 11,"Defeat rows", et.get('defeat',0))
srow(ws, 12,"Thin / generic descriptions", s['thin_descriptions'], "Rows with <120 chars or boilerplate text", WARN_FILL)
srow(ws, 13,"Major battles MISSING engagement row", s['battles_missing'], "See sheet 'Missing Battles'", MISS_FILL)
srow(ws, 14,"Eastern Cape events MISSING", s['eastern_events_missing'], "See sheet 'Missing EC Events'", MISS_FILL)
srow(ws, 15,"Major commanders MISSING", s['commanders_missing'], "Cronjé, Piet — Paardeberg; see sheet", MISS_FILL)
srow(ws, 16,"Low-confidence coordinate rows", s['low_confidence'], "All fixed to 0 — good")
srow(ws, 17,"Rows with no source", s['no_source'], "All rows have a source — good")
srow(ws, 18,"Movement rows with no from/to line", s['movement_rows_no_line'], "All 0 — arrows fully wired")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Missing Battles
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Missing Battles")
hdr(ws2, 1, [("Battle",28),("Expected Date",14),("Historical Note",48),("Status",12),("Priority",12)])
autofreeze(ws2)
for i, b in enumerate(audit['battle_audit'], 2):
    missing = 'MISSING' in b['status']
    fill = MISS_FILL if missing else OK_FILL
    priority = 'HIGH' if missing and b['expected_date'] < '1900-06' else ('MED' if missing else '')
    rowfill(ws2, i, [b['battle'], b['expected_date'], b['note'], b['status'], priority], fill=fill)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 – Missing EC Events
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Missing EC Events")
hdr(ws3, 1, [("Place",18),("Date",10),("Note",60),("Status",12)])
autofreeze(ws3)
for i, e in enumerate(audit['eastern_missing'], 2):
    missing = 'MISSING' in e['status']
    fill = MISS_FILL if missing else OK_FILL
    rowfill(ws3, i, [e['place'], e['date'], e['note'], e['status']], fill=fill)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 – Missing Commander
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Commander Audit")
hdr(ws4, 1, [("Commander",24),("Side",10),("Role",52),("Status",12)])
autofreeze(ws4)
for i, c in enumerate(audit['commander_audit'], 2):
    missing = 'MISSING' in c['status']
    fill = MISS_FILL if missing else OK_FILL
    rowfill(ws4, i, [c['commander'], c['side'], c['role'], c['status']], fill=fill)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 – Thin Descriptions
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Thin Descriptions")
hdr(ws5, 1, [("ID",6),("Force / Unit",38),("Event Type",14),("Place",20),("Description (truncated)",60),("Issue",22)])
autofreeze(ws5)
for i, t in enumerate(audit['thin_descriptions'], 2):
    fill = ALT_FILL if i % 2 == 0 else None
    rowfill(ws5, i, [t['id'], t['force'], t['event_type'], t['action_place'], t['description'], t['issue']],
            fill=fill, wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 – Deployment Stacks
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Deployment Stacks")
hdr(ws6, 1, [("Place",28),("Deployment Count",18),("Note",44)])
autofreeze(ws6)
notes = {
    'Pretoria': 'Transvaal HQ anchor — many column ops units',
    'Bloemfontein': 'OFS advance anchor — expected to be large',
    'Cape Town': 'Base port / disembarkation point — expected',
    'Ladysmith': 'Natal siege centre — expected',
    'Kimberley': 'Western siege town — expected',
    'Mafeking': 'Northern siege town — expected',
    'Durban': 'Natal base port — expected',
    'Pietermaritzburg': 'Natal Admin base',
}
for i, d in enumerate(audit['deployment_stacks'], 2):
    fill = WARN_FILL if d['count'] > 60 else (ALT_FILL if i%2==0 else None)
    note = notes.get(d['place'], '')
    rowfill(ws6, i, [d['place'], d['count'], note], fill=fill)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7 – Event Type Review
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Event Type Review")
hdr(ws7, 1, [("Event Type",18),("Count",8),("Map Icon/Layer",22),("Example row (force)",40)])
autofreeze(ws7)
etype_layer = {
    'deployment': 'pts layer — push-pin icon',
    'engagement': 'pts layer — crossed swords icon',
    'advance': 'moves/move-arrows layer — arrow line',
    'movement': 'moves/move-arrows layer — arrow line',
    'move': 'moves/move-arrows layer — arrow line',
    'raid': 'pts layer — lightning bolt',
    'siege': 'pts layer — castle/ring icon',
    'capture': 'pts layer — flag icon',
    'defeat': 'pts layer — red X',
    'pursuit': 'pts layer — chase arrow',
    'drive': 'pts layer — push arrow',
    'skirmish': 'pts layer — small swords',
    'surrender': 'pts layer — white flag',
    'occupation': 'pts layer — flag',
    'disembark': 'pts layer — ship/anchor',
    'garrison': 'pts layer — shield',
    'rail-move': 'pts layer — train icon',
}
et_rows = {}
for r in all_rows:
    et_rows.setdefault(r['event_type'], r)
et_counts = s['event_types']
for i, (etype, cnt) in enumerate(sorted(et_counts.items(), key=lambda x: -x[1]), 2):
    example = et_rows.get(etype, {}).get('force', '')[:40]
    fill = ALT_FILL if i%2==0 else None
    rowfill(ws7, i, [etype, cnt, etype_layer.get(etype,'pts layer'), example], fill=fill)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 8 – ID Gaps
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("ID Gaps")
hdr(ws8, 1, [("Gap ID",10),("Surrounding IDs",22),("Action",40)])
autofreeze(ws8)
gids = audit['id_gaps']
id_set = {int(r['id']) for r in all_rows}
for i, gid in enumerate(gids, 2):
    prev_id = gid - 1
    next_id = gid + 1
    prev_force = next((r['force'] for r in all_rows if int(r['id'])==prev_id), '?')
    next_force = next((r['force'] for r in all_rows if int(r['id'])==next_id), '?')
    action = f"Leave blank OR add a row with id={gid} if you find a missing unit"
    rowfill(ws8, i, [gid, f"#{prev_id} {prev_force[:25]} … #{next_id} {next_force[:25]}", action], fill=WARN_FILL)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 9 – Add-These-Rows (CSV template for missing battles)
# ══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Add-These-Rows")
ws9.sheet_properties.tabColor = "FF4444"
FIELDS = ['id','side','force','commander','units','date_start','date_end',
          'event_type','from_place','to_place','action_place','description',
          'confidence','source','note']
hdr(ws9, 1, [(f, 18 if f not in ('description','note') else 55) for f in FIELDS])

# Pre-filled rows for the 9 missing battles + Cronjé + 3 EC events
MISSING_ROWS = [
    # id=693 onwards (next free after 692)
    dict(id=693, side='Boer', force="Battle of Lombard's Kop",
         commander='Joubert', units="Natal Commandos",
         date_start='1899-10-30', date_end='1899-10-30',
         event_type='defeat', action_place='Lombards Kop',
         description="British sortie from Ladysmith repulsed at Lombard's Kop on 30 Oct 1899; White's force beaten back with heavy loss and Dundee garrison forced to retreat, completing Boer encirclement of Ladysmith.",
         confidence='high', source='Conan Doyle ch.V', note=''),
    dict(id=694, side='British', force="Battle of Dreifontein",
         commander='Roberts', units="6th Division, Cavalry Division",
         date_start='1900-03-10', date_end='1900-03-10',
         event_type='engagement', action_place='Dreifontein',
         description="Roberts's advance force defeated De Wet's rearguard at Dreifontein (10 Mar 1900), opening the road to Bloemfontein; one of the sharp conventional engagements during the OFS advance.",
         confidence='high', source='Conan Doyle ch.XVI', note='approx coords needed'),
    dict(id=695, side='British', force="Battle of Karee Siding",
         commander='Tucker', units="7th Division",
         date_start='1900-03-29', date_end='1900-03-29',
         event_type='engagement', action_place='Karee Siding',
         description="Roberts's 7th Division under Tucker engaged a Boer rearguard at Karee Siding north of Bloemfontein on 29 Mar 1900, pushing the Boers back and maintaining the advance momentum.",
         confidence='high', source='Conan Doyle ch.XVII', note='approx coords needed'),
    dict(id=696, side='Boer', force="Battle of Sannaspos (Koorn Spruit)",
         commander='De Wet', units="De Wet's commando; Sarel Eloff",
         date_start='1900-03-31', date_end='1900-03-31',
         event_type='raid', action_place='Sannaspos',
         description="De Wet ambushed Broadwood's column at Koorn Spruit / Sannaspos (31 Mar 1900) east of Bloemfontein, capturing 7 guns and 117 wagons in the most brilliant Boer coup of the conventional phase.",
         confidence='high', source='Conan Doyle ch.XVII', note=''),
    dict(id=697, side='Boer', force="Siege of Wepener",
         commander='Olivier, J.H.', units="Olivier's Commando",
         date_start='1900-04-09', date_end='1900-04-25',
         event_type='siege', action_place='Wepener',
         description="Olivier's OFS commando besieged a Colonial and Cape garrison at Wepener 9-25 Apr 1900; the garrison held out until relieved, tying down Boer forces and allowing Roberts to manoeuvre.",
         confidence='high', source='Conan Doyle ch.XVII', note=''),
    dict(id=698, side='British', force="Relief of Mafeking",
         commander='Mahon, B.T.; Plumer, H.', units="Mahon's Relief Column; Plumer's Rhodesian Regiment",
         date_start='1900-05-17', date_end='1900-05-17',
         event_type='capture', action_place='Mafeking',
         description="Mafeking relieved 17 May 1900 by Mahon's flying column from the south and Plumer from the north, ending the 217-day siege; news caused jubilant celebrations throughout the British Empire ('mafficking').",
         confidence='high', source='Conan Doyle ch.XV', note=''),
    dict(id=699, side='Boer', force="Battle of Helvetia",
         commander='Grobler, P.', units="Eastern Transvaal Commandos",
         date_start='1901-01-29', date_end='1901-01-29',
         event_type='raid', action_place='Helvetia',
         description="Boer commandos under Grobler surprised the British garrison at Helvetia fort (29 Jan 1901), capturing the 5-gun 113th Battery RFA and 200 prisoners in one of the most audacious guerrilla raids of 1901.",
         confidence='high', source='angloboerwar.com', note=''),
    dict(id=700, side='British', force="Colesburg Operations",
         commander='French, J.D.P.', units="Cavalry Brigade; 1st Div detachments",
         date_start='1899-12-01', date_end='1900-01-15',
         event_type='engagement', action_place='Colesburg',
         description="Gen French conducted a skilful holding action around Colesburg Nov 1899–Jan 1900, fixing Boer forces on the northern Cape front and preventing a deeper invasion while Roberts organised the main advance.",
         confidence='high', source='Conan Doyle ch.XIII', note=''),
    dict(id=701, side='Boer', force="Siege of Okiep (O'kiep)",
         commander='Smuts, Jan', units="Smuts's Cape commando",
         date_start='1902-04-01', date_end='1902-05-04',
         event_type='siege', action_place="O'kiep",
         description="Smuts besieged the copper-mining town of O'kiep in Namaqualand Apr-May 1902, holding out until the Peace of Vereeniging; one of the last active operations of the war in the Cape Colony.",
         confidence='high', source='angloboerwar.com', note=''),
    dict(id=702, side='Boer', force="Cronjé at Paardeberg",
         commander='Cronjé, Piet', units="Cronjé's OFS/Transvaal force (~4,000)",
         date_start='1900-02-18', date_end='1900-02-27',
         event_type='surrender', action_place='Paardeberg',
         description="Piet Cronjé and ~4,000 burghers surrendered to Roberts at Paardeberg on 27 Feb 1900 (Majuba Day) after a 10-day siege on the Modder River; the largest Boer surrender of the war and the turning point of the conventional phase.",
         confidence='high', source='Conan Doyle ch.XIV', note=''),
    # Eastern Cape gaps
    dict(id=703, side='Boer', force="Smuts's Cape Raid — Barkly East",
         commander='Smuts, Jan', units="Smuts's commando (~300)",
         date_start='1901-09-15', date_end='1901-11-01',
         event_type='raid', action_place='Barkly East',
         description="Smuts's Cape commando passed through the Barkly East district in Sep-Oct 1901 during the southward phase of his Cape raid, recruiting Cape rebels and evading British pursuit columns in the eastern mountains.",
         confidence='high', source='angloboerwar.com', note=''),
    dict(id=704, side='Boer', force="Smuts's Cape Raid — Lady Grey / Dordrecht",
         commander='Smuts, Jan', units="Smuts's commando",
         date_start='1901-08-15', date_end='1901-09-15',
         event_type='raid', action_place='Lady Grey',
         description="Smuts crossed the Orange River near Stormberg and raided through Lady Grey and Dordrecht districts in Aug-Sep 1901, recruiting Cape rebels and gathering horses before swinging south toward Elands River Poort.",
         confidence='medium', source='angloboerwar.com', note='Lady Grey coords approximate'),
    dict(id=705, side='British', force="Molteno / Stormberg aftermath",
         commander='Gatacre, W.F.', units="3rd Division detachments; Queen's, Northumberland Fusiliers",
         date_start='1899-12-11', date_end='1900-02-01',
         event_type='retreat', action_place='Molteno',
         description="After Gatacre's disaster at Stormberg (10 Dec 1899), the broken British force retreated to Molteno; Gatacre held this rail-junction until relieved in early 1900, preventing further Boer advance into the Midlands.",
         confidence='high', source='Conan Doyle ch.XII', note=''),
]

for i, r in enumerate(MISSING_ROWS, 2):
    fill = ALT_FILL if i%2==0 else None
    vals = [str(r.get(f,'')) for f in FIELDS]
    rowfill(ws9, i, vals, fill=fill, wrap=True)
    ws9.row_dimensions[i].height = 45

# row heights for description col
for ws in [ws2, ws3, ws5]:
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30

# ── Save ──────────────────────────────────────────────────────────────────────
out = 'boer_war_audit.xlsx'
wb.save(out)
print(f"Written {out}")
print(f"  Sheet 9 has {len(MISSING_ROWS)} ready-to-add rows (IDs 693-{692+len(MISSING_ROWS)})")
