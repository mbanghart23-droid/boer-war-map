"""
Check N: Undocumented Location Transitions.

For each regiment, sort all events by date. Where the place changes significantly
(>20 km between consecutive events), check whether a movement-type row exists
that documents HOW the unit got there (advance, retreat, rail_move, redeployment,
raid, movement, drive, pursuit).

If no such row exists â†’ flag as an undocumented transition.

This is different from the travel-speed check (A):
  - Check A: "is this physically impossible?"
  - Check N: "is this transition documented, regardless of whether it's possible?"

A transition can be perfectly plausible (unit railed 400 km) but still be a gap
in the historical record if we haven't added the connecting row.
"""
import json, csv, math, re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent.parent

rows   = list(csv.DictReader(open(HERE/"data"/"movements.csv", encoding="utf-8")))
gj     = json.load(open(HERE/"docs"/"data"/"events.geojson", encoding="utf-8"))

row_by_id  = {r["id"]: r for r in rows}
feats      = [f for f in gj["features"]]
event_feats= {f["properties"]["id"]: f for f in feats if f["properties"]["kind"]=="event"}
move_feats = [f for f in feats if f["properties"]["kind"] in ("move","route")]

def parse_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], "%Y-%m-%d")
    except: return None

def haversine(c1, c2):
    lon1,lat1 = math.radians(c1[0]),math.radians(c1[1])
    lon2,lat2 = math.radians(c2[0]),math.radians(c2[1])
    dlat=lat2-lat1; dlon=lon2-lon1
    a=math.sin(dlat/2)**2+math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 6371*2*math.asin(math.sqrt(a))

MOVE_TYPES = {"advance","retreat","movement","raid","pursuit","drive",
              "redeployment","rail_move","rail-move","disembark"}

# â”€â”€ Build per-regiment event timeline from GeoJSON â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
reg_timeline = defaultdict(list)
for fid, feat in event_feats.items():
    p = feat["properties"]
    d = parse_date(p.get("date_start",""))
    if not d: continue
    for g in (p.get("groups") or []):
        reg = g[0]
        reg_timeline[reg].append({
            "id": fid,
            "date": d,
            "date_end": parse_date(p.get("date_end","")),
            "place": p.get("place",""),
            "coords": feat["geometry"]["coordinates"],
            "force": p.get("force",""),
            "side": p.get("side",""),
            "event_type": p.get("event_type",""),
        })

# â”€â”€ Build per-regiment movement row index â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# For each regiment, collect all rows from CSV that are movement-type
reg_move_rows = defaultdict(list)
for r in rows:
    if r.get("event_type","") not in MOVE_TYPES: continue
    # Associate to regiment via units field (same logic as build_map.py)
    # Use force name as a loose match â€” not perfect but avoids re-implementing canon
    reg_move_rows[r.get("force","")].append(r)

# Also index move features by force name
move_feat_by_force = defaultdict(list)
for f in move_feats:
    p = f["properties"]
    move_feat_by_force[p.get("force","")].append(f)

# â”€â”€ Main check â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
MIN_DISTANCE_KM = 25   # ignore place changes smaller than this (coord noise / nearby towns)
issues = []

for reg, timeline in sorted(reg_timeline.items()):
    timeline.sort(key=lambda e: e["date"])

    for i in range(len(timeline)-1):
        a = timeline[i]
        b = timeline[i+1]

        # How far apart are the two locations?
        dist = haversine(a["coords"], b["coords"])
        if dist < MIN_DISTANCE_KM:
            continue   # same area, no transition needed

        # What's the time window between these events?
        t_start = a.get("date_end") or a["date"]
        t_end   = b["date"]
        days    = max((t_end - t_start).days, 0)

        # Is there a movement-type ROW (CSV) for this regiment that:
        #   1. Has a date_start/date_end that overlaps the gap window
        #   2. Is movement-type
        # We match loosely by regiment name appearing in force or units field
        reg_lower = reg.lower()

        has_move_row = False
        bridging_rows = []

        for r in rows:
            if r.get("event_type","") not in MOVE_TYPES:
                continue
            rd = parse_date(r.get("date_start",""))
            if not rd: continue

            # Does this row fall in the gap window?
            in_window = (t_start - datetime(1,1,1).__class__.min if True else False)
            # Simpler: row date must be between a's end and b's start (with 30d slop)
            from datetime import timedelta
            if not (t_start - timedelta(days=30) <= rd <= t_end + timedelta(days=30)):
                continue

            # Does this row involve our regiment?
            force_lower = r.get("force","").lower()
            units_lower = r.get("units","").lower()
            # Match: regiment name appears in force or units
            if reg_lower in force_lower or reg_lower in units_lower:
                has_move_row = True
                bridging_rows.append(r["id"])
                break

            # Also try partial match (e.g. "17th lancers" matches "17th (Duke of Cambridge's Own) Lancers")
            reg_words = [w for w in reg_lower.split() if len(w) > 4]
            if reg_words and all(w in force_lower or w in units_lower for w in reg_words):
                has_move_row = True
                bridging_rows.append(r["id"])
                break

        # Also check if there's a move FEATURE (line) in GeoJSON for this force
        if not has_move_row:
            for force_key, mfeats in move_feat_by_force.items():
                fk_lower = force_key.lower()
                reg_words = [w for w in reg_lower.split() if len(w) > 4]
                if not reg_words: continue
                if not (reg_lower in fk_lower or all(w in fk_lower for w in reg_words)):
                    continue
                for mf in mfeats:
                    mp = mf["properties"]
                    md = parse_date(mp.get("date_start",""))
                    if md and t_start - __import__('datetime').timedelta(days=30) <= md <= t_end + __import__('datetime').timedelta(days=30):
                        has_move_row = True
                        break
                if has_move_row:
                    break

        if not has_move_row:
            # Categorise severity by distance and time gap
            if dist > 300 and days < 30:
                sev = "HIGH"
            elif dist > 150:
                sev = "HIGH"
            elif dist > 75:
                sev = "MEDIUM"
            else:
                sev = "LOW"

            # Suggest what the missing row should look like
            # Guess the movement type
            if days <= 3 and dist > 100:
                likely_type = "rail_move (fast over long distance)"
            elif a.get("side") == "Boer":
                likely_type = "raid or redeployment"
            else:
                likely_type = "advance or redeployment"

            issues.append({
                "severity": sev,
                "regiment": reg,
                "side": a.get("side",""),
                "from_id": a["id"],
                "from_date": a["date"].strftime("%Y-%m-%d"),
                "from_place": a["place"],
                "from_force": a["force"],
                "to_id": b["id"],
                "to_date": b["date"].strftime("%Y-%m-%d"),
                "to_place": b["place"],
                "to_force": b["force"],
                "dist_km": round(dist),
                "gap_days": days,
                "likely_type": likely_type,
                "note": f"{dist:.0f}km in {days}d â€” no movement row documents this transition. "
                        f"Missing row should be event_type='{likely_type}' "
                        f"from_place='{a['place']}' to_place='{b['place']}' "
                        f"date ~{a['date'].strftime('%Y-%m')}.",
            })

# â”€â”€ Summary â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from collections import Counter
sev_count = Counter(i["severity"] for i in issues)
print(f"UNDOCUMENTED TRANSITIONS: {len(issues)} total")
for s in ["HIGH","MEDIUM","LOW"]:
    print(f"  {s}: {sev_count[s]}")

# Show a sample
print("\nSample HIGH issues:")
for iss in sorted(issues, key=lambda x: -x["dist_km"]):
    if iss["severity"] != "HIGH": continue
    print(f"  {iss['regiment']} ({iss['side']})")
    print(f"    {iss['from_date']} @ {iss['from_place']} â†’ {iss['to_date']} @ {iss['to_place']}")
    print(f"    {iss['dist_km']}km / {iss['gap_days']}d | suggest: {iss['likely_type']}")
    print()

# â”€â”€ Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Undocumented Transitions"

HDR = ["Severity","Regiment","Side","Dist km","Gap days",
       "From ID","From Date","From Place","From Force",
       "To ID","To Date","To Place","To Force",
       "Likely Move Type","Research Note"]
ws.append(HDR)

hf = PatternFill("solid", fgColor="0d1b2a")
hfn = Font(bold=True, color="FFFFFF", size=10)
for c in range(1, len(HDR)+1):
    ws.cell(1,c).fill = hf; ws.cell(1,c).font = hfn

SEV_COL = {"HIGH":"FF6600","MEDIUM":"FFAA00","LOW":"CCCCCC"}

for iss in sorted(issues, key=lambda x: (["HIGH","MEDIUM","LOW"].index(x["severity"]), -x["dist_km"])):
    ws.append([
        iss["severity"], iss["regiment"], iss["side"],
        iss["dist_km"], iss["gap_days"],
        iss["from_id"], iss["from_date"], iss["from_place"], iss["from_force"],
        iss["to_id"],   iss["to_date"],   iss["to_place"],   iss["to_force"],
        iss["likely_type"], iss["note"],
    ])
    r = ws.max_row
    col = SEV_COL.get(iss["severity"],"CCCCCC")
    ws.cell(r,1).fill = PatternFill("solid", fgColor=col)
    ws.cell(r,1).font = Font(bold=True, color="FFFFFF" if iss["severity"]=="HIGH" else "333333")
    ws.cell(r,4).alignment = Alignment(horizontal="center")
    ws.cell(r,5).alignment = Alignment(horizontal="center")
    ws.cell(r,15).alignment = Alignment(wrap_text=True)

WIDTHS = [10,28,8,9,9,8,12,22,32,8,12,22,32,22,70]
for i,w in enumerate(WIDTHS,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# â”€â”€ Sheet 2: by regiment (easier for a researcher working on one unit) â”€â”€â”€â”€â”€â”€â”€â”€â”€
ws2 = wb.create_sheet("By Regiment")
ws2.append(["Regiment","Side","# Gaps","Total km undocumented","Worst gap (km)","Worst gap route"])
hf2=PatternFill("solid",fgColor="1a3a00"); hfn2=Font(bold=True,color="90FF90",size=10)
for c in range(1,7): ws2.cell(1,c).fill=hf2; ws2.cell(1,c).font=hfn2

by_reg = defaultdict(list)
for iss in issues:
    by_reg[iss["regiment"]].append(iss)

for reg, iss_list in sorted(by_reg.items(), key=lambda x: -sum(i["dist_km"] for i in x[1])):
    worst = max(iss_list, key=lambda i: i["dist_km"])
    total_km = sum(i["dist_km"] for i in iss_list)
    ws2.append([
        reg,
        iss_list[0]["side"],
        len(iss_list),
        total_km,
        worst["dist_km"],
        f"{worst['from_place']} â†’ {worst['to_place']} ({worst['from_date']}â†’{worst['to_date']})"
    ])

for col,w in zip("ABCDEF",[30,8,8,18,14,55]):
    ws2.column_dimensions[col].width=w
ws2.freeze_panes="A2"

out = HERE/"tools"/"undocumented_transitions.xlsx"
wb.save(out)
print(f"\nSaved: {out}")

