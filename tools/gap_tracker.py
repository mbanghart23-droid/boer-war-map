"""
Comprehensive Gap Tracker — unified research document combining:
  1. Chronological gaps (>30 days, same or different location)
  2. Undocumented location transitions (new place, no movement row)
  3. Canonical unit merging (fragmented Boer group names → one unit)

Output: gap_tracker.xlsx
  Sheet: All Gaps          — every gap sorted by priority
  Sheet: By Unit           — one section per unit, full timeline + gaps highlighted
  Sheet: Boer Units        — Boer-only summary, sorted by gap severity
  Sheet: British Units     — British-only summary
  Sheet: Research Status   — actionable tracker with Status / Researcher / Notes columns
"""

import csv, json, math, re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent.parent
rows_csv = list(csv.DictReader(open(HERE / "data" / "movements.csv", encoding="utf-8")))
gj       = json.load(open(HERE / "docs" / "data" / "events.geojson", encoding="utf-8"))

# ── Canonical unit mapping ─────────────────────────────────────────────────────
# Maps ANY group name → canonical unit name
# Groups not listed here keep their own name
CANONICAL = {
    # Smuts
    "Smuts commando":                           "Smuts commando",
    "Smuts' Cape commando":                     "Smuts commando",
    "Smuts/van Deventer commando":              "Smuts commando",
    "Smuts's Cape Raid — departure":            "Smuts commando",
    "Smuts's Cape Raid — Elands River Poort":   "Smuts commando",
    "Smuts's Cape Raid — spread through Namaqualand": "Smuts commando",
    "Smuts's Raid � Barkly East sector":   "Smuts commando",
    "Smuts's Raid � Lady Grey / Dordrecht":"Smuts commando",
    # Kritzinger
    "Kritzinger commando":                      "Kritzinger commando",
    "Kritzinger commando (1st invasion)":       "Kritzinger commando",
    "Kritzinger commando (2nd invasion)":       "Kritzinger commando",
    "Kritzinger and Scheepers commandos":       "Kritzinger commando",
    "Kritzinger's Cape raid � entry":      "Kritzinger commando",
    "Kritzinger's Cape raid � south":      "Kritzinger commando",
    # Scheepers
    "Scheepers commando":                       "Scheepers commando",
    "Scheepers & Fouch\xe9 commandos":          "Scheepers commando",
    # De Wet
    "De Wet's Commando":                        "De Wet's commando",
    "De Wet's first escape — Roodeval":         "De Wet's commando",
    "De Wet's first invasion of Cape Colony":   "De Wet's commando",
    "De Wet's second escape — February 1901":   "De Wet's commando",
    # Botha
    "Botha's commando":                         "Botha's commando",
    "Botha's Natal raiding force":              "Botha's commando",
    "Botha's Commando (eastern Transvaal)":     "Botha's commando",
    "Botha's Natal offensive":                  "Botha's commando",
    # De la Rey
    "De la Rey's commando":                     "De la Rey's commando",
    "Delarey's Commando":                       "De la Rey's commando",
    "De la Rey and Kemp commando":              "De la Rey's commando",
    # OFS / Prinsloo
    "OFS commando":                             "OFS commando",
    "OFS commandos":                            "OFS commando",
    "OFS commandos (Rouxville & Thaba Nchu)":  "OFS commando",
    "Prinsloo's force — trapped in Brandwater Basin": "OFS commando",
    "Prinsloo's OFS force":                    "OFS commando",
    "Prinsloo's Commando (Jacob Prinsloo)":    "OFS commando",
    "Olivier's Commando":                       "OFS commando",
    "Olivier (OFS commandos)":                 "OFS commando",
    # Hertzog
    "Hertzog's western Cape raid":              "Hertzog's commando",
    # Van Reenen
    "van Reenen commando":                      "van Reenen commando",
    "van Reenen commando (under Gen Malan)":    "van Reenen commando",
    "Cape-rebel exiles (van Reenen's men)":     "van Reenen commando",
    # Malan
    "Malan commando":                           "Malan commando",
    "Malan commando (return attempt)":          "Malan commando",
    "Malan's Commando (Cape Rebels)":           "Malan commando",
    # Van Deventer
    "Van Deventer commando":                    "Van Deventer commando",
    # Wessels
    "Wessels commando":                         "Wessels commando",
    "Wessels' commando":                        "Wessels commando",
    # Natal
    "Natal commandos":                          "Natal commandos",
    "Natal commando":                           "Natal commandos",
    "Boer invasion of Natal":                   "Natal commandos",
    "Boer siege of Ladysmith — investment":     "Natal commandos",
    "Joubert's raid south of Ladysmith":        "Natal commandos",
    # Cronje / Paardeberg
    "Cronje's army":                            "Cronje's army",
    "Cronj� surrenders at Paardeberg":     "Cronje's army",
    # Western Transvaal
    "Western Transvaal commandos":              "Western Transvaal commandos",
    "Western Transvaal commando":               "Western Transvaal commandos",
    # British Royal Artillery catch-all
    "Royal Artillery":                          "Royal Artillery (composite)",
    "Royal Garrison Artillery":                 "Royal Artillery (composite)",
    "Royal Horse Artillery":                    "Royal Artillery (composite)",
    # British MI catch-all
    "Mounted Infantry":                         "Mounted Infantry (composite)",
    # Imperial Yeomanry
    "Imperial Yeomanry":                        "Imperial Yeomanry",
    # Lotter
    "Lotter commando":                          "L\xf6tter commando",
    "L\xf6tter commando":                       "L\xf6tter commando",
    "Lotter (capture & execution)":             "L\xf6tter commando",
}

def canonical(g):
    return CANONICAL.get(g, g)

# ── Structural composite units ─────────────────────────────────────────────────
# These canonical names represent a SINGLE CSV label that actually covers
# multiple independent units (batteries, battalions, contingents) operating
# simultaneously at different locations.  Their "HIGH" gaps are data-model
# artefacts — not research gaps — because the same label appears in two places
# at once.  We assign severity = "STRUCTURAL" so they never inflate the HIGH
# count and appear in their own section in the workbook.
STRUCTURAL_COMPOSITES = {
    # Already named "(composite)" via CANONICAL above
    "Royal Artillery (composite)",
    "Mounted Infantry (composite)",
    # NZ contingents now split by number in build_map.py REGT_RULES;
    # "New Zealand Mounted Infantry" remains as fallback for unspecified NZ rows
    "New Zealand Mounted Infantry",
    # Cape Police split into 3 groups via build_map.py REGT_RULES
    # Natal Carbineers removed: single regiment
    # Town Guard split by garrison town (Mafeking/Kimberley/Ladysmith/generic)
    "Yorkshire Regiment",   # composite: 1st Bn (Natal) vs 2nd Bn (western Transvaal) vs KOYLI
    "Royal Garrison Artillery",       # single placeholder row
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def parse_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], "%Y-%m-%d")
    except: return None

def haversine(c1, c2):
    lon1,lat1 = math.radians(c1[0]),math.radians(c1[1])
    lon2,lat2 = math.radians(c2[0]),math.radians(c2[1])
    dlat=lat2-lat1; dlon=lon2-lon1
    a=math.sin(dlat/2)**2+math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 6371*2*math.asin(math.sqrt(a))

MOVE_TYPES = {"advance","retreat","movement","raid","pursuit","drive",
              "redeployment","rail_move","rail-move","disembark"}

# ── Build per-canonical-unit event timelines ──────────────────────────────────
event_feats = {f["properties"]["id"]: f
               for f in gj["features"] if f["properties"]["kind"]=="event"}
move_feats  = [f for f in gj["features"] if f["properties"]["kind"] in ("move","route")]

unit_timeline = defaultdict(list)
for fid, feat in event_feats.items():
    p = feat["properties"]
    d = parse_date(p.get("date_start",""))
    if not d: continue
    for g in (p.get("groups") or []):
        cu = canonical(g[0])
        unit_timeline[cu].append({
            "id": fid,
            "date": d,
            "date_end": parse_date(p.get("date_end","")),
            "place": p.get("place",""),
            "coords": feat["geometry"]["coordinates"],
            "force": p.get("force",""),
            "side": p.get("side",""),
            "event_type": p.get("event_type",""),
            "description": p.get("description","")[:120],
            "source": p.get("source",""),
            "confidence": p.get("confidence",""),
            "orig_group": g[0],
        })

# ── Build movement-row lookup (CSV rows that are movement-type) ───────────────
csv_move_rows = [r for r in rows_csv if r.get("event_type","") in MOVE_TYPES]

def has_movement_bridge(cu, t_start, t_end, side):
    """Return True if there is a movement-type row for this unit in the window."""
    cu_lower = cu.lower()
    cu_words = [w for w in cu_lower.split() if len(w) > 4]
    window_start = t_start - timedelta(days=60)
    window_end   = t_end   + timedelta(days=60)
    for r in csv_move_rows:
        rd = parse_date(r.get("date_start",""))
        if not rd or not (window_start <= rd <= window_end):
            continue
        force_lower = r.get("force","").lower()
        units_lower = r.get("units","").lower()
        combined = force_lower + " " + units_lower
        if cu_lower in combined:
            return True
        if cu_words and all(w in combined for w in cu_words):
            return True
    # Also check move GeoJSON features
    for mf in move_feats:
        mp = mf["properties"]
        md = parse_date(mp.get("date_start",""))
        if not md or not (window_start <= md <= window_end):
            continue
        mf_lower = mp.get("force","").lower()
        if cu_lower in mf_lower:
            return True
        if cu_words and all(w in mf_lower for w in cu_words):
            return True
    return False

# ── Source suggestions by unit ────────────────────────────────────────────────
SOURCE_HINTS = {
    "17th Lancers":            "Regimental history '17th (Duke of Cambridge's Own) Lancers'; Haig Papers (NLS); Conan Doyle chs XIV-XIX",
    "9th Lancers":             "Regimental history '9th (Queen's Royal) Lancers 1715-1936'; Amery Times History Vol 3-5",
    "Cape Mounted Rifles":     "CMR regimental history; SAMAD archives; Warwick 'Black People and the South African War'",
    "Brabant's Horse":         "Benyon 'Proconsul and Paramountcy'; Nasson 'The South African War'",
    "Smuts commando":          "Hancock 'Smuts Vol 1'; Smuts 'Memoirs of the Boer War'; Kruger 'Goodbye Dolly Gray' ch 30",
    "Kritzinger commando":     "Nasson 'Abraham Esau's War'; Grundlingh 'The Dynamics of Treason'; De Wet 'Three Years War'",
    "Scheepers commando":      "Van der Waag; Nasson 'Abraham Esau's War'; SAMAD commando records",
    "De Wet's commando":       "De Wet 'Three Years War'; Pakenham 'The Boer War'; Amery Times History Vol 4",
    "OFS commando":            "Prinsloo 'Die Brandwag'; Amery Times History Vol 4; SANDF Documentation Centre",
    "Botha's commando":        "Meintjes 'Louis Botha'; Pakenham 'The Boer War'; Amery Times History",
    "De la Rey's commando":    "Meintjes 'De la Rey — Lion of the West'; Amery Times History Vol 5-6",
    "Lotter commando":         "Nasson 'Abraham Esau's War'; trial records (Graaff-Reinet))",
    "Malan commando":          "Nasson 'Abraham Esau's War'; Cape rebel records SAMAD",
    "South African Light Horse":"Conan Doyle; Amery Times History Vol 4; SALH regimental records",
    "Bethune's MI":            "Amery Times History; regimental diary (NAUK WO95 series)",
    "Cape Police":             "Rauch 'The History of the Cape Police'; SAMAD Cape Police records",
    "Imperial Yeomanry":       "Doyle 'The Great Boer War'; Amery; IY regimental diaries (NAUK WO95)",
    "New Zealand Mounted Infantry": "Powles 'The History of the Canterbury Mounted Rifles'; ANZ archives",
    "Free State commando":     "De Wet 'Three Years War'; OFS State Archives (VAB)",
    "Hertzog's commando":      "Nienaber 'Hertzog'; Pienaar 'With Steyn and De Wet'; SANDF Documentation Centre",
}

def source_hint(cu):
    for key, hint in SOURCE_HINTS.items():
        if key.lower() in cu.lower() or cu.lower() in key.lower():
            return hint
    return "Check: NAUK WO108 (Boer War records); SAMAD archives; Amery 'Times History of the War in South Africa'"

# ── Analyse each unit ─────────────────────────────────────────────────────────
GAP_THRESHOLD   = 30    # days — minimum time gap to flag
DIST_THRESHOLD  = 25    # km — minimum distance to flag as location change

all_gaps = []

for cu, timeline in sorted(unit_timeline.items()):
    timeline.sort(key=lambda e: e["date"])
    side = timeline[0].get("side","") if timeline else ""

    for i in range(len(timeline)-1):
        a = timeline[i]
        b = timeline[i+1]

        t_from  = a.get("date_end") or a["date"]
        t_to    = b["date"]
        gap_days = max((t_to - t_from).days, 0)
        dist_km  = haversine(a["coords"], b["coords"])

        is_time_gap       = gap_days >= GAP_THRESHOLD
        is_location_change= dist_km  >= DIST_THRESHOLD

        if not (is_time_gap or is_location_change):
            continue

        # Check for bridging movement row
        bridge = has_movement_bridge(cu, t_from, t_to, side)

        # Gap type
        if is_location_change and not bridge:
            gap_type = "Location change — no movement row"
        elif is_time_gap and is_location_change and bridge:
            gap_type = "Time gap + location (movement row exists)"
        elif is_time_gap and not is_location_change:
            gap_type = "Time gap (same location)"
        elif is_time_gap and is_location_change and not bridge:
            gap_type = "Time gap + location change — no movement row"
        else:
            gap_type = "Location change (movement row exists)"

        # Skip if movement row exists AND time gap is <30d AND location changes moderately
        # Exception: always keep STRUCTURAL gaps so they are documented in the workbook
        if bridge and gap_days < GAP_THRESHOLD and dist_km < 200 and cu not in STRUCTURAL_COMPOSITES:
            continue

        # Severity
        if cu in STRUCTURAL_COMPOSITES:
            sev = "STRUCTURAL"  # data-model artefact, not a resolvable research gap
        elif not bridge and (dist_km > 200 or (gap_days > 90 and dist_km > DIST_THRESHOLD)):
            sev = "HIGH"       # genuinely missing movement documentation (must involve location change)
        elif gap_days > 60 or (not bridge and dist_km > 75):
            sev = "MEDIUM"     # includes bridged long gaps and minor unbridged moves
        else:
            sev = "LOW"

        # Suggest movement type
        if sev == "STRUCTURAL":
            suggested = "N/A — split composite unit label in CSV first"
        elif dist_km > 100 and gap_days <= 3:
            suggested = "rail_move"
        elif side == "Boer":
            suggested = "redeployment / raid"
        elif gap_days > 180:
            suggested = "redeployment (long garrison / column work)"
        else:
            suggested = "advance / redeployment"

        all_gaps.append({
            "unit":         cu,
            "side":         side,
            "severity":     sev,
            "gap_type":     gap_type,
            "bridge_exists":bridge,
            "gap_days":     gap_days,
            "dist_km":      round(dist_km),
            "from_id":      a["id"],
            "from_date":    a["date"].strftime("%Y-%m-%d"),
            "from_date_end":(a["date_end"].strftime("%Y-%m-%d") if a.get("date_end") else ""),
            "from_place":   a["place"],
            "from_force":   a["force"],
            "from_event":   a["event_type"],
            "to_id":        b["id"],
            "to_date":      b["date"].strftime("%Y-%m-%d"),
            "to_place":     b["place"],
            "to_force":     b["force"],
            "to_event":     b["event_type"],
            "suggested_type": suggested,
            "sources":      source_hint(cu),
            "status":       "Cannot resolve" if cu in STRUCTURAL_COMPOSITES else "Not started",
        "notes_auto":   f"STRUCTURAL: '{cu}' is a composite label covering multiple independent units simultaneously. Cannot be resolved without splitting into separate unit rows (e.g. by battery, battalion, or contingent number)." if cu in STRUCTURAL_COMPOSITES else "",
            "researcher":   "",
            "resolved_date":"",
            "resolution":   "",
            "notes":        "",
        })

# Totals
from collections import Counter
resolvable = [g for g in all_gaps if g["severity"] != "STRUCTURAL"]
structural = [g for g in all_gaps if g["severity"] == "STRUCTURAL"]
print(f"Total gaps: {len(all_gaps)}  (structural noise excluded from counts below)")
sev_c = Counter(g["severity"] for g in resolvable)
side_c= Counter(g["side"] for g in resolvable)
for s in ["HIGH","MEDIUM","LOW"]: print(f"  {s}: {sev_c[s]}")
print(f"  British: {side_c['British']}  Boer: {side_c['Boer']}")
print(f"  STRUCTURAL (composite-label artefacts, not resolvable without data split): {len(structural)}")
struct_units = Counter(g["unit"] for g in structural)
for u, n in sorted(struct_units.items(), key=lambda x: -x[1]):
    print(f"    {u}: {n} gaps")
print()
print("Boer unit gap summary:")
boer_units = defaultdict(list)
for g in all_gaps:
    if g["side"]=="Boer": boer_units[g["unit"]].append(g)
for u, gs in sorted(boer_units.items(), key=lambda x: (-sum(1 for g in x[1] if g["severity"]=="HIGH"), x[0])):
    hi = sum(1 for g in gs if g["severity"]=="HIGH")
    print(f"  {u}: {len(gs)} gaps ({hi} HIGH)")

# ── Excel ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

SEV_COL  = {"HIGH":"FF6600","MEDIUM":"FFAA00","LOW":"DDDDDD","STRUCTURAL":"9966CC"}
SEV_ORD  = {"HIGH":0,"MEDIUM":1,"LOW":2,"STRUCTURAL":3}
HDR_DARK = PatternFill("solid", fgColor="0d1b2a")
HDR_BOER = PatternFill("solid", fgColor="2a1500")
HDR_BRIT = PatternFill("solid", fgColor="001533")
HDR_RES  = PatternFill("solid", fgColor="0a2a0a")
WF = Font(bold=True, color="FFFFFF", size=10)

def write_hdr(ws, headers, fill, font=None):
    ws.append(headers)
    f = font or WF
    for c in range(1, len(headers)+1):
        ws.cell(1,c).fill = fill
        ws.cell(1,c).font = f
        ws.cell(1,c).alignment = Alignment(wrap_text=True)

def sev_fill(sev):
    col = SEV_COL.get(sev,"CCCCCC")
    return PatternFill("solid", fgColor=col)

def sev_font(sev):
    return Font(bold=True, color="FFFFFF" if sev in ("HIGH","MEDIUM","STRUCTURAL") else "333333", size=10)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: All Gaps
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "All Gaps"
HDR1 = ["Severity","Unit","Side","Gap Type","Gap Days","Dist km","Bridge?",
        "From ID","From Date","From Place","From Force","From Event",
        "To ID","To Date","To Place","To Force","To Event",
        "Suggested Move Type","Sources"]
write_hdr(ws1, HDR1, HDR_DARK)

for g in sorted(all_gaps, key=lambda x:(SEV_ORD[x["severity"]], x["unit"], x["from_date"])):
    ws1.append([
        g["severity"], g["unit"], g["side"], g["gap_type"],
        g["gap_days"], g["dist_km"],
        "YES" if g["bridge_exists"] else "NO",
        g["from_id"], g["from_date"], g["from_place"], g["from_force"], g["from_event"],
        g["to_id"],   g["to_date"],   g["to_place"],   g["to_force"],   g["to_event"],
        g["suggested_type"], g["sources"],
    ])
    r = ws1.max_row
    ws1.cell(r,1).fill = sev_fill(g["severity"])
    ws1.cell(r,1).font = sev_font(g["severity"])
    ws1.cell(r,7).font = Font(bold=True, color="CC0000" if g["bridge_exists"] is False else "008800")
    ws1.cell(r,19).alignment = Alignment(wrap_text=True)

WIDTHS1=[10,26,8,28,9,8,7,7,11,22,30,14,7,11,22,30,14,20,60]
for i,w in enumerate(WIDTHS1,1): ws1.column_dimensions[get_column_letter(i)].width=w
ws1.freeze_panes="A2"; ws1.auto_filter.ref=ws1.dimensions

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Research Status Tracker
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Research Status")
HDR2 = ["Status","Severity","Unit","Side","Gap Type","Gap Days","Dist km",
        "From Date","From Place","To Date","To Place",
        "Suggested Move Type","Sources to Check",
        "Researcher","Date Resolved","Resolution / New Row IDs","Notes"]
write_hdr(ws2, HDR2, HDR_RES)

STATUS_FILL = {
    "Not started": PatternFill("solid", fgColor="FFE0E0"),
    "In progress": PatternFill("solid", fgColor="FFFACC"),
    "Resolved":    PatternFill("solid", fgColor="E0FFE0"),
    "Cannot resolve": PatternFill("solid", fgColor="E0E0E0"),
}

for g in sorted(all_gaps, key=lambda x:(SEV_ORD[x["severity"]], x["unit"], x["from_date"])):
    ws2.append([
        g["status"], g["severity"], g["unit"], g["side"], g["gap_type"],
        g["gap_days"], g["dist_km"],
        g["from_date"], g["from_place"],
        g["to_date"],   g["to_place"],
        g["suggested_type"], g["sources"],
        g["researcher"], g["resolved_date"], g["resolution"],
        g.get("notes_auto","") or g["notes"],
    ])
    r = ws2.max_row
    ws2.cell(r,1).fill = STATUS_FILL.get(g["status"], STATUS_FILL["Not started"])
    ws2.cell(r,2).fill = sev_fill(g["severity"])
    ws2.cell(r,2).font = sev_font(g["severity"])
    ws2.cell(r,13).alignment = Alignment(wrap_text=True)
    ws2.cell(r,17).alignment = Alignment(wrap_text=True)

WIDTHS2=[13,10,26,8,28,9,8,11,22,11,22,20,60,14,13,22,45]
for i,w in enumerate(WIDTHS2,1): ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A2"; ws2.auto_filter.ref=ws2.dimensions

# Add data validation for Status column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(
    type="list",
    formula1='"Not started,In progress,Resolved,Cannot resolve"',
    showDropDown=False
)
ws2.add_data_validation(dv)
dv.add(f"A2:A{ws2.max_row+200}")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Boer Units
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Boer Units")
HDR3 = ["Unit","# Gaps","HIGH","MEDIUM","LOW","STRUCTURAL","No Bridge","Total Gap Days",
        "Worst Gap","Worst Gap Route","Key Sources"]
write_hdr(ws3, HDR3, HDR_BOER)

boer_summary = defaultdict(list)
for g in all_gaps:
    if g["side"]=="Boer": boer_summary[g["unit"]].append(g)

for cu, gs in sorted(boer_summary.items(),
                     key=lambda x:(-sum(1 for g in x[1] if not g["bridge_exists"] and g["severity"]=="HIGH"), x[0])):
    hi   = sum(1 for g in gs if g["severity"]=="HIGH")
    med  = sum(1 for g in gs if g["severity"]=="MEDIUM")
    lo   = sum(1 for g in gs if g["severity"]=="LOW")
    strc = sum(1 for g in gs if g["severity"]=="STRUCTURAL")
    nobr = sum(1 for g in gs if not g["bridge_exists"])
    tot  = sum(g["gap_days"] for g in gs)
    worst= max(gs, key=lambda g: g["gap_days"])
    ws3.append([
        cu, len(gs), hi, med, lo, strc, nobr, tot, worst["gap_days"],
        f"{worst['from_place']} -> {worst['to_place']} ({worst['from_date']} to {worst['to_date']})",
        source_hint(cu),
    ])
    r = ws3.max_row
    if hi > 0:
        ws3.cell(r,3).fill = sev_fill("HIGH"); ws3.cell(r,3).font = sev_font("HIGH")
    if med > 0:
        ws3.cell(r,4).fill = sev_fill("MEDIUM"); ws3.cell(r,4).font = sev_font("MEDIUM")
    if strc > 0:
        ws3.cell(r,6).fill = sev_fill("STRUCTURAL"); ws3.cell(r,6).font = sev_font("STRUCTURAL")
    ws3.cell(r,11).alignment = Alignment(wrap_text=True)

WIDTHS3=[28,7,7,7,7,10,8,13,10,55,65]
for i,w in enumerate(WIDTHS3,1): ws3.column_dimensions[get_column_letter(i)].width=w
ws3.freeze_panes="A2"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: British Units
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("British Units")
write_hdr(ws4, HDR3, HDR_BRIT)

brit_summary = defaultdict(list)
for g in all_gaps:
    if g["side"]=="British": brit_summary[g["unit"]].append(g)

for cu, gs in sorted(brit_summary.items(),
                     key=lambda x:(-sum(1 for g in x[1] if not g["bridge_exists"] and g["severity"]=="HIGH"), x[0])):
    hi   = sum(1 for g in gs if g["severity"]=="HIGH")
    med  = sum(1 for g in gs if g["severity"]=="MEDIUM")
    lo   = sum(1 for g in gs if g["severity"]=="LOW")
    strc = sum(1 for g in gs if g["severity"]=="STRUCTURAL")
    nobr = sum(1 for g in gs if not g["bridge_exists"])
    tot  = sum(g["gap_days"] for g in gs)
    worst= max(gs, key=lambda g: g["gap_days"])
    ws4.append([
        cu, len(gs), hi, med, lo, strc, nobr, tot, worst["gap_days"],
        f"{worst['from_place']} -> {worst['to_place']} ({worst['from_date']} to {worst['to_date']})",
        source_hint(cu),
    ])
    r = ws4.max_row
    if hi > 0:
        ws4.cell(r,3).fill = sev_fill("HIGH"); ws4.cell(r,3).font = sev_font("HIGH")
    if med > 0:
        ws4.cell(r,4).fill = sev_fill("MEDIUM"); ws4.cell(r,4).font = sev_font("MEDIUM")
    if strc > 0:
        ws4.cell(r,6).fill = sev_fill("STRUCTURAL"); ws4.cell(r,6).font = sev_font("STRUCTURAL")
    ws4.cell(r,11).alignment = Alignment(wrap_text=True)

for i,w in enumerate(WIDTHS3,1): ws4.column_dimensions[get_column_letter(i)].width=w
ws4.freeze_panes="A2"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Per-unit full timeline (shows each unit's events + gaps inline)
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Unit Timelines")
HDR5=["Unit","Side","Date Start","Date End","Place","Force / Column",
      "Event Type","Description","Source","<<GAP>>","Gap Days","Dist km","Gap Type"]
write_hdr(ws5, HDR5, HDR_DARK)

GAP_FILL = PatternFill("solid", fgColor="330000")
GAP_FONT = Font(bold=True, color="FF8888", size=10)
EVT_BRIT = PatternFill("solid", fgColor="00102a")
EVT_BOER = PatternFill("solid", fgColor="1a0a00")

# Build gap lookup: (cu, from_id) -> gap info
gap_lookup = {}
for g in all_gaps:
    gap_lookup[(g["unit"], g["from_id"])] = g

printed_units = set()
for cu, timeline in sorted(unit_timeline.items()):
    timeline.sort(key=lambda e: e["date"])
    if not timeline: continue
    side = timeline[0].get("side","")

    # Unit header row
    ws5.append([cu, side] + [""]*11)
    r = ws5.max_row
    hdr_fill = HDR_BRIT if side=="British" else HDR_BOER
    for c in range(1,14):
        ws5.cell(r,c).fill = hdr_fill
        ws5.cell(r,c).font = Font(bold=True, color="FFFFFF", size=11)
    ws5.row_dimensions[r].height = 16

    for i, evt in enumerate(timeline):
        # Event row
        ws5.append([
            "", "",
            evt["date"].strftime("%Y-%m-%d"),
            evt["date_end"].strftime("%Y-%m-%d") if evt.get("date_end") else "",
            evt["place"], evt["force"], evt["event_type"],
            evt["description"][:100], evt["source"][:50],
            "", "", "", "",
        ])
        r = ws5.max_row
        ef = EVT_BRIT if side=="British" else EVT_BOER
        for c in range(1,14): ws5.cell(r,c).fill = ef
        ws5.cell(r,8).alignment = Alignment(wrap_text=True)

        # Gap row after this event?
        gap = gap_lookup.get((cu, evt["id"]))
        if gap:
            ws5.append([
                "", "", "", "", "", "", "", "", "",
                f"GAP: {gap['gap_type']}",
                gap["gap_days"], gap["dist_km"], gap["gap_type"],
            ])
            r = ws5.max_row
            for c in range(1,14): ws5.cell(r,c).fill = GAP_FILL
            ws5.cell(r,10).font = GAP_FONT
            ws5.cell(r,11).font = Font(bold=True, color="FF8888")
            ws5.cell(r,12).font = Font(bold=True, color="FF8888")

    # Spacer
    ws5.append([""] * 13)

WIDTHS5=[26,8,11,11,22,30,14,60,40,28,9,8,28]
for i,w in enumerate(WIDTHS5,1): ws5.column_dimensions[get_column_letter(i)].width=w
ws5.freeze_panes="A2"

out = HERE/"tools"/"gap_tracker.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print(f"Sheets: All Gaps | Research Status | Boer Units | British Units | Unit Timelines")
