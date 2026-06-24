"""
unit_stats.py — compute per-unit statistics from movements.csv + casualties + strength.

Outputs:
  data/unit_stats.json   — full stat object, consumed by map/API
  tools/unit_stats.xlsx  — Excel workbook with multiple sheets for review

Metrics computed:
  COMPUTED (from movements.csv + gazetteer):
    - days_in_theatre         first-to-last event span
    - engagement_count        number of event_type='engagement' rows
    - km_total                sum of straight-line from→to distances
    - km_advance              km during Roberts advance phase (≤ 1900-09-30)
    - km_guerrilla            km during guerrilla phase (> 1900-09-30)
    - locations_count         distinct action_places visited
    - engagements_per_month   engagement_count / (days_in_theatre / 30.44)
    - phase_days_*            days in each war phase with events
    - first_event / last_event
    - commanders              list of distinct commanders (from 'commander' col)

  DATA-DRIVEN (from unit_casualties.csv + unit_strength.csv):
    - total_killed / wounded / captured / missing / casualties
    - peak_strength           highest strength recorded
    - strength_at_start       strength nearest first event date
    - cas_rate                casualties / peak_strength (%)

  VALIDATION FLAGS:
    - has_strength_data       bool
    - has_casualty_data       bool
    - data_completeness       0-100 score based on filled fields
"""
import csv, json, math, datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL = True
except ImportError:
    EXCEL = False
    print("openpyxl not found — Excel output skipped")

ROOT = Path(__file__).parent.parent
CSV_PATH   = ROOT / "data" / "movements.csv"
GAZ_PATH   = ROOT / "data" / "gazetteer.json"
CAS_PATH   = ROOT / "data" / "unit_casualties.csv"
STR_PATH   = ROOT / "data" / "unit_strength.csv"
JSON_OUT   = ROOT / "data" / "unit_stats.json"
XLSX_OUT   = Path(__file__).parent / "unit_stats.xlsx"

# ── War phases ────────────────────────────────────────────────────────────────
PHASE_BORDERS = [
    ("invasion",  datetime.date(1899, 10, 11), datetime.date(1900,  2, 10)),
    ("advance",   datetime.date(1900,  2, 11), datetime.date(1900,  9, 30)),
    ("guerrilla", datetime.date(1900, 10,  1), datetime.date(1902,  5, 31)),
]
def phase_of(d):
    for name, s, e in PHASE_BORDERS:
        if s <= d <= e: return name
    return "other"

# ── Helpers ───────────────────────────────────────────────────────────────────
def pd(s):
    try: return datetime.date.fromisoformat(s.strip())
    except: return None

def haversine(a, b):
    """Great-circle distance in km between [lat,lon] pairs."""
    if not a or not b: return 0.0
    lat1,lon1 = math.radians(a[0]), math.radians(a[1])
    lat2,lon2 = math.radians(b[0]), math.radians(b[1])
    dlat, dlon = lat2-lat1, lon2-lon1
    h = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 2*6371*math.asin(math.sqrt(h))

# ── Load data ─────────────────────────────────────────────────────────────────
gaz   = json.load(open(GAZ_PATH))
coord = {k: v[:2] for k, v in gaz.items()}   # place → [lat, lon]

mov_rows = list(csv.DictReader(open(CSV_PATH, encoding="utf-8")))
cas_rows = list(csv.DictReader(open(CAS_PATH, encoding="utf-8")))
str_rows = list(csv.DictReader(open(STR_PATH, encoding="utf-8")))

# Index by force
by_force = defaultdict(list)
for r in mov_rows: by_force[r["force"]].append(r)

by_force_cas = defaultdict(list)
for r in cas_rows: by_force_cas[r["force"]].append(r)

by_force_str = defaultdict(list)
for r in str_rows: by_force_str[r["force"]].append(r)

def safe_int(v):
    try: return int(str(v).replace(",","").strip())
    except: return 0

# ── Compute stats per force ───────────────────────────────────────────────────
all_stats = {}

for force, rows in by_force.items():
    dates = [pd(r["date_start"]) for r in rows if pd(r["date_start"])]
    if not dates:
        continue
    dates.sort()
    first, last = dates[0], dates[-1]
    days_in_theatre = (last - first).days + 1

    # Engagements
    eng_rows = [r for r in rows if r["event_type"] == "engagement"]
    engagement_count = len(eng_rows)

    # Engagements per phase
    eng_by_phase = defaultdict(int)
    for r in eng_rows:
        d = pd(r["date_start"])
        if d: eng_by_phase[phase_of(d)] += 1

    # Distance
    km_total = km_adv = km_guer = 0.0
    for r in rows:
        fp, tp = r.get("from_place",""), r.get("to_place","") or r.get("action_place","")
        if fp and tp and fp != tp:
            dist = haversine(coord.get(fp), coord.get(tp))
            km_total += dist
            d = pd(r["date_start"])
            if d:
                ph = phase_of(d)
                if ph == "advance":   km_adv  += dist
                if ph == "guerrilla": km_guer += dist

    # Days per phase
    phase_dates = defaultdict(set)
    for r in rows:
        d = pd(r["date_start"])
        if d: phase_dates[phase_of(d)].add(d)
    phase_days = {ph: len(ds) for ph, ds in phase_dates.items()}

    # Locations
    locations = set(r["action_place"] for r in rows if r.get("action_place"))

    # Commanders
    commanders = sorted(set(
        r["commander"].strip() for r in rows
        if r.get("commander","").strip() and r["commander"].strip() != ""
    ))

    # Engagements per month
    eng_per_month = round(engagement_count / max(days_in_theatre / 30.44, 1), 3)

    # Side
    side = rows[0]["side"] if rows else ""

    # ── Casualty data ─────────────────────────────────────────────────────────
    cas = by_force_cas.get(force, [])
    total_killed    = sum(safe_int(r["killed"])   for r in cas)
    total_wounded   = sum(safe_int(r["wounded"])  for r in cas)
    total_captured  = sum(safe_int(r["captured"]) for r in cas)
    total_missing   = sum(safe_int(r["missing"])  for r in cas)
    total_casualties = sum(safe_int(r["total"]) for r in cas) or (
        total_killed + total_wounded + total_captured + total_missing)

    cas_events = [{"date": r["date"], "event": r["event"],
                   "killed": safe_int(r["killed"]), "wounded": safe_int(r["wounded"]),
                   "captured": safe_int(r["captured"]), "missing": safe_int(r["missing"]),
                   "total": safe_int(r["total"]), "source": r["source"]}
                  for r in cas]

    # ── Strength data ─────────────────────────────────────────────────────────
    strs = sorted(by_force_str.get(force, []), key=lambda r: r["date"])
    peak_strength = max((safe_int(r["total"]) for r in strs), default=0)
    strength_at_start = 0
    if strs:
        nearest = min(strs, key=lambda r: abs((pd(r["date"]) or first) - first))
        strength_at_start = safe_int(nearest["total"])

    str_timeline = [{"date": r["date"], "total": safe_int(r["total"]),
                     "officers": safe_int(r["officers"]),
                     "other_ranks": safe_int(r["other_ranks"]),
                     "horses": safe_int(r["horses"]),
                     "guns": safe_int(r["guns"]),
                     "source": r["source"]}
                    for r in strs]

    # ── Attrition rate ────────────────────────────────────────────────────────
    cas_rate = round(total_casualties / peak_strength * 100, 1) if peak_strength else None

    # ── Data completeness score ───────────────────────────────────────────────
    score = 0
    if engagement_count > 0: score += 20
    if km_total > 0:         score += 20
    if cas:                  score += 30
    if strs:                 score += 20
    if commanders:           score += 10
    completeness = score

    all_stats[force] = {
        # identity
        "force": force,
        "side": side,
        # time
        "first_event": str(first),
        "last_event":  str(last),
        "days_in_theatre": days_in_theatre,
        "phase_days": phase_days,
        # movement
        "km_total":     round(km_total, 1),
        "km_advance":   round(km_adv, 1),
        "km_guerrilla": round(km_guer, 1),
        "locations_count": len(locations),
        "locations": sorted(locations),
        # engagements
        "engagement_count": engagement_count,
        "engagements_per_month": eng_per_month,
        "engagements_by_phase": dict(eng_by_phase),
        # people
        "commanders": commanders,
        "peak_strength":      peak_strength,
        "strength_at_start":  strength_at_start,
        "strength_timeline":  str_timeline,
        # casualties
        "total_killed":     total_killed,
        "total_wounded":    total_wounded,
        "total_captured":   total_captured,
        "total_missing":    total_missing,
        "total_casualties": total_casualties,
        "casualty_rate_pct": cas_rate,
        "casualty_events":  cas_events,
        # meta
        "has_casualty_data": bool(cas),
        "has_strength_data": bool(strs),
        "data_completeness": completeness,
    }

# ── Write JSON ────────────────────────────────────────────────────────────────
json.dump(all_stats, open(JSON_OUT, "w", encoding="utf-8"), indent=2)
print("JSON written: %s  (%d units)" % (JSON_OUT, len(all_stats)))

# ── Write Excel ───────────────────────────────────────────────────────────────
if not EXCEL:
    print("Skipping Excel (no openpyxl)")
else:
    wb = openpyxl.Workbook()

    HDR = Font(bold=True)
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    YELLOW = PatternFill("solid", fgColor="FFFACD")
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")

    def freeze_and_filter(ws, row=1):
        ws.freeze_panes = ws.cell(row+1, 1)
        ws.auto_filter.ref = ws.dimensions

    def col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    COLS1 = ["Force", "Side", "First Event", "Last Event", "Days",
             "Engagements", "Eng/Month", "Km Total", "Km Advance", "Km Guerrilla",
             "Locations", "Peak Strength", "Total Casualties",
             "Killed", "Wounded", "Captured", "Cas Rate %",
             "Has Cas Data", "Has Str Data", "Completeness"]
    ws.append(COLS1)
    for c in ws[1]: c.font = HDR; c.fill = GRAY

    stats_list = sorted(all_stats.values(), key=lambda s: s["side"] + s["force"])
    for s in stats_list:
        ws.append([
            s["force"], s["side"],
            s["first_event"], s["last_event"], s["days_in_theatre"],
            s["engagement_count"], s["engagements_per_month"],
            s["km_total"], s["km_advance"], s["km_guerrilla"],
            s["locations_count"], s["peak_strength"], s["total_casualties"],
            s["total_killed"], s["total_wounded"], s["total_captured"],
            s["casualty_rate_pct"],
            "Y" if s["has_casualty_data"] else "",
            "Y" if s["has_strength_data"] else "",
            s["data_completeness"],
        ])
    col_widths(ws, [45,8,12,12,6,8,8,8,8,8,8,10,10,7,7,8,8,8,8,8])
    freeze_and_filter(ws)

    # ── Sheet 2: Casualties ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Casualties")
    COLS2 = ["Force", "Date", "Event", "Killed", "Wounded", "Captured", "Missing", "Total", "Source"]
    ws2.append(COLS2)
    for c in ws2[1]: c.font = HDR; c.fill = GRAY
    for r in cas_rows:
        ws2.append([r["force"], r["date"], r["event"],
                    safe_int(r["killed"]), safe_int(r["wounded"]),
                    safe_int(r["captured"]), safe_int(r["missing"]),
                    safe_int(r["total"]), r["source"]])
    col_widths(ws2, [45,12,30,7,7,8,7,7,50])
    freeze_and_filter(ws2)

    # ── Sheet 3: Strength ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Strength")
    COLS3 = ["Force", "Date", "Officers", "Other Ranks", "Total", "Horses", "Guns", "Source", "Note"]
    ws3.append(COLS3)
    for c in ws3[1]: c.font = HDR; c.fill = GRAY
    for r in str_rows:
        ws3.append([r["force"], r["date"],
                    safe_int(r["officers"]), safe_int(r["other_ranks"]),
                    safe_int(r["total"]), safe_int(r["horses"]), safe_int(r["guns"]),
                    r["source"], r.get("note","")])
    col_widths(ws3, [45,12,9,10,8,8,6,50,40])
    freeze_and_filter(ws3)

    # ── Sheet 4: Engagement Audit ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Engagement Audit")
    COLS4 = ["Force", "Side", "Date", "Place", "Description", "Confidence",
             "Has Casualty Data", "Source"]
    ws4.append(COLS4)
    for c in ws4[1]: c.font = HDR; c.fill = GRAY
    cas_index = set((r["force"], r["date"]) for r in cas_rows)
    for r in sorted(mov_rows, key=lambda x: x["date_start"]):
        if r["event_type"] != "engagement": continue
        key = (r["force"], r["date_start"])
        has_cas = "Y" if key in cas_index else ""
        row = [r["force"], r["side"], r["date_start"], r["action_place"],
               r["description"][:120], r["confidence"], has_cas, r["source"][:60]]
        ws4.append(row)
        if not has_cas:
            ws4.cell(ws4.max_row, 7).fill = YELLOW
    col_widths(ws4, [45,8,12,25,80,10,10,50])
    freeze_and_filter(ws4)

    # ── Sheet 5: Distance Leaders ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Distance Leaders")
    ws5.append(["Force", "Side", "Km Total", "Km Advance", "Km Guerrilla",
                "Locations", "Days"])
    for c in ws5[1]: c.font = HDR; c.fill = GRAY
    top = sorted(stats_list, key=lambda s: s["km_total"], reverse=True)[:100]
    for s in top:
        ws5.append([s["force"], s["side"], s["km_total"], s["km_advance"],
                    s["km_guerrilla"], s["locations_count"], s["days_in_theatre"]])
    col_widths(ws5, [45,8,9,9,10,9,7])
    freeze_and_filter(ws5)

    # ── Sheet 6: Coverage Gaps (engagements without casualty data) ────────────
    ws6 = wb.create_sheet("Cas Data Gaps")
    ws6.append(["Force", "Side", "Engagement Date", "Place", "Confidence"])
    for c in ws6[1]: c.font = HDR; c.fill = GRAY
    for r in sorted(mov_rows, key=lambda x: x["date_start"]):
        if r["event_type"] != "engagement": continue
        if (r["force"], r["date_start"]) not in cas_index:
            ws6.append([r["force"], r["side"], r["date_start"],
                        r["action_place"], r["confidence"]])
    col_widths(ws6, [45,8,12,25,10])
    freeze_and_filter(ws6)

    wb.save(XLSX_OUT)
    print("Excel written: %s" % XLSX_OUT)

# ── Summary to console ────────────────────────────────────────────────────────
total_eng = sum(s["engagement_count"] for s in all_stats.values())
has_cas   = sum(1 for s in all_stats.values() if s["has_casualty_data"])
has_str   = sum(1 for s in all_stats.values() if s["has_strength_data"])
total_km  = sum(s["km_total"] for s in all_stats.values())
top5_km   = sorted(all_stats.values(), key=lambda s: s["km_total"], reverse=True)[:5]
top5_eng  = sorted(all_stats.values(), key=lambda s: s["engagement_count"], reverse=True)[:5]

print("\n-- SUMMARY ------------------------------------------------------")
print("Units tracked:       %d" % len(all_stats))
print("Total engagements:   %d" % total_eng)
print("Total km computed:   %.0f" % total_km)
print("Units w/ cas data:   %d" % has_cas)
print("Units w/ str data:   %d" % has_str)
print("\nTop 5 by km travelled:")
for s in top5_km:
    print("  %-45s  %.0f km" % (s["force"][:45], s["km_total"]))
print("\nTop 5 by engagement count:")
for s in top5_eng:
    print("  %-45s  %d engagements" % (s["force"][:45], s["engagement_count"]))
