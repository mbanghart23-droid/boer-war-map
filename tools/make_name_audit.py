"""
make_name_audit.py
==================
Boer War Eastern Cape — Name Quality Audit
Run from the boer-war-eastern-cape directory:
    python tools/make_name_audit.py

Produces: boer_war_name_audit.xlsx (7 audit sheets)

Requires: openpyxl, csv (stdlib), json (stdlib), re (stdlib), difflib (stdlib)
Optional but recommended: pip install rapidfuzz  (falls back to difflib if absent)
"""

import csv
import json
import re
import sys
import os
from collections import defaultdict, Counter
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Try rapidfuzz for faster/better fuzzy matching; fall back to difflib
# ---------------------------------------------------------------------------
try:
    from rapidfuzz import fuzz as _fuzz
    def fuzzy_ratio(a, b):
        return _fuzz.token_set_ratio(a.lower(), b.lower())
except ImportError:
    from difflib import SequenceMatcher
    def fuzzy_ratio(a, b):
        return int(SequenceMatcher(None, a.lower(), b.lower()).ratio() * 100)

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# PATHS (relative to CWD = boer-war-eastern-cape)
# ---------------------------------------------------------------------------
CSV_PATH  = os.path.join("data", "movements.csv")
JSON_PATH = os.path.join("data", "unit_roster.json")
OUT_PATH  = "boer_war_name_audit.xlsx"

# ---------------------------------------------------------------------------
# STYLE CONSTANTS
# ---------------------------------------------------------------------------
FILL_HEADER  = PatternFill("solid", fgColor="1F3864")   # dark navy
FILL_BAD     = PatternFill("solid", fgColor="FFD7D7")   # light red
FILL_OK      = PatternFill("solid", fgColor="D7FFD7")   # light green
FILL_WARN    = PatternFill("solid", fgColor="FFFACD")   # light yellow
FILL_ALT     = PatternFill("solid", fgColor="F2F5FB")   # blue-grey alternating

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_NORMAL  = Font(name="Calibri", size=10)
FONT_BOLD    = Font(name="Calibri", bold=True, size=10)

ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")

# ---------------------------------------------------------------------------
# QSA CLASP LIST
# ---------------------------------------------------------------------------
QSA_CLASPS = [
    "CAPE COLONY", "NATAL", "RHODESIA",
    "DEFENCE OF KIMBERLEY", "TALANA", "ELANDSLAAGTE",
    "DEFENCE OF LADYSMITH", "BELMONT", "MODDER RIVER",
    "TUGELA HEIGHTS", "RELIEF OF KIMBERLEY", "PAARDEBERG",
    "RELIEF OF LADYSMITH", "DRIEFONTEIN", "WEPENER",
    "DEFENCE OF MAFEKING", "JOHANNESBURG", "DIAMOND HILL",
    "WITTEBERGEN", "BELFAST", "TRANSVAAL", "ORANGE FREE STATE",
    "LAING'S NEK", "RELIEF OF MAFEKING",
    "SOUTH AFRICA 1901", "SOUTH AFRICA 1902",
]

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def load_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def load_roster(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    # Return list of unit name strings
    return [entry["unit"] for entry in data if "unit" in entry]


def style_header_row(ws, headers, col_widths=None):
    """Write bold navy header row and set column widths."""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill  = FILL_HEADER
        cell.font  = FONT_HEADER
        cell.alignment = ALIGN_CENTER
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"


def fill_row(ws, row_num, values, fill=None):
    """Write a data row; apply fill if given, else alternate."""
    chosen_fill = fill if fill else (FILL_ALT if row_num % 2 == 0 else None)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_WRAP
        if chosen_fill:
            cell.fill = chosen_fill


def parse_date(s):
    """Parse YYYY-MM-DD into a date object; return None on failure."""
    if not s or not s.strip():
        return None
    s = s.strip()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def is_valid_date_fmt(s):
    """Return True if s is exactly YYYY-MM-DD format."""
    if not s or not s.strip():
        return False
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", s.strip()))


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print(f"Loading {CSV_PATH} …")
    rows = load_csv(CSV_PATH)
    print(f"  {len(rows)} rows loaded.")

    print(f"Loading {JSON_PATH} …")
    roster_names = load_roster(JSON_PATH)
    roster_lower = {n.lower(): n for n in roster_names}   # lower→official
    print(f"  {len(roster_names)} roster entries loaded.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # =========================================================================
    # SHEET 1 — Unit Name Mismatches
    # =========================================================================
    print("Building Sheet 1: Unit Name Mismatches …")
    ws1 = wb.create_sheet("1 Unit Name Mismatches")

    headers1 = ["Row ID", "Side", "force (raw)", "Best Roster Match", "Match %",
                "Issue Type", "Notes"]
    widths1  = [8, 8, 40, 40, 10, 28, 40]
    style_header_row(ws1, headers1, widths1)

    # Broad-province / country words that appear in force names but aren't specific units
    broad_terms = re.compile(
        r"\b(OFS|Orange Free State|Cape Colony|Transvaal|Natal|South Africa|"
        r"Royal Field Artillery|Royal Horse Artillery|Royal Garrison Artillery)\b"
    )

    row_num = 2
    for r in rows:
        row_id   = r.get("id", "")
        side     = r.get("side", "")
        force    = r.get("force", "").strip()
        if not force:
            continue

        issues = []

        # ---- capitalisation check ----
        if force != force.title() and force == force.lower():
            issues.append("All-lowercase")

        # ---- fuzzy match against roster ----
        best_name  = ""
        best_score = 0
        for official in roster_names:
            s = fuzzy_ratio(force, official)
            if s > best_score:
                best_score = s
                best_name  = official

        # ---- exact lower match (case discrepancy) ----
        if force.lower() in roster_lower and force != roster_lower[force.lower()]:
            issues.append("Case mismatch with roster")

        # ---- abbreviation detection (Bn vs Battalion, Bty vs Battery, etc.) ----
        abbrev_map = {
            r"\bBn\b":       "Battalion",
            r"\bBty\b":      "Battery",
            r"\bBde\b":      "Brigade",
            r"\bRegt\b":     "Regiment",
            r"\bHvy\b":      "Heavy",
            r"\bMt\b":       "Mounted",
            r"\bInf\b":      "Infantry",
            r"\bRFA\b":      "Royal Field Artillery",
            r"\bRHA\b":      "Royal Horse Artillery",
            r"\bRGA\b":      "Royal Garrison Artillery",
            r"\bCMR\b":      "Cape Mounted Rifles",
            r"\bIY\b":       "Imperial Yeomanry",
            r"\bMI\b":       "Mounted Infantry",
            r"\bDMT\b":      "District Mounted Troops",
        }
        for pat, expansion in abbrev_map.items():
            if re.search(pat, force):
                issues.append(f"Abbreviation: '{re.search(pat,force).group()}' vs '{expansion}'")

        # ---- missing number when roster expects one (e.g. RFA battery number) ----
        if re.search(r"Royal Field Artillery", force, re.I) and not re.search(r"\d+(st|nd|rd|th)?\s+Batt?e?ry", force, re.I):
            issues.append("RFA mentioned without specific battery number")
        if re.search(r"Royal Horse Artillery", force, re.I) and not re.search(r"\b[A-Z]\b Battery|\b\d+", force):
            issues.append("RHA mentioned without battery identifier")

        # ---- no match in roster ----
        if best_score < 80 and side.lower() == "british":
            issues.append(f"No close roster match (<80%)")

        if not issues:
            fill = FILL_OK
            issue_str = "OK"
        elif "No close roster match" in " ".join(issues):
            fill = FILL_BAD
            issue_str = "; ".join(issues)
        else:
            fill = FILL_WARN
            issue_str = "; ".join(issues)

        fill_row(ws1, row_num,
                 [row_id, side, force, best_name, best_score, issue_str, ""],
                 fill=fill)
        row_num += 1

    # =========================================================================
    # SHEET 2 — Commander Name Format Audit
    # =========================================================================
    print("Building Sheet 2: Commander Name Format Audit …")
    ws2 = wb.create_sheet("2 Commander Names")

    headers2 = ["Row ID", "event_type", "commander (raw)", "Issue Type", "Suggestion"]
    widths2  = [8, 16, 40, 42, 40]
    style_header_row(ws2, headers2, widths2)

    # Event types that must have a commander
    must_have_commander = {
        "engagement", "battle", "raid", "advance", "defeat", "capture",
        "skirmish", "drive", "pursuit"
    }

    row_num = 2
    for r in rows:
        row_id   = r.get("id", "")
        evt      = r.get("event_type", "").strip().lower()
        cmd      = r.get("commander", "").strip()

        issues = []
        suggestion = ""

        if not cmd:
            if evt in must_have_commander:
                issues.append("Blank commander on engagement/combat row")
            else:
                # non-combat blank — record as OK but light note
                fill_row(ws2, row_num,
                         [row_id, evt, "(blank)", "OK — non-combat row blank", ""],
                         fill=FILL_OK)
                row_num += 1
                continue
        else:
            # ---- check for comma (Surname, Firstname format) ----
            if "," not in cmd:
                issues.append("No comma — expected 'Surname, Initials/Firstname' format")
                # Try to suggest inversion: if last word looks like a surname
                parts = cmd.split()
                if len(parts) >= 2:
                    # Heuristic: last token is surname
                    surname  = parts[-1]
                    rest     = " ".join(parts[:-1])
                    suggestion = f"{surname}, {rest}"

            # ---- check for initials-first format (J.D.P. French) ----
            # Pattern: starts with one or more X. sequences then a word
            if re.match(r"^([A-Z]\.){1,5}\s+[A-Z][a-z]", cmd):
                issues.append("Initials-first: should be 'Surname, Initials' format")
                parts = cmd.rsplit(" ", 1)
                if len(parts) == 2:
                    suggestion = f"{parts[1]}, {parts[0]}"

            # ---- double period ----
            if ".." in cmd:
                issues.append("Double period in initials")
                suggestion = cmd.replace("..", ".")

            # ---- no periods after initials that look like initials ----
            # Single uppercase letters not followed by . (and not at end)
            if re.search(r"\b[A-Z]\b[^.]", cmd):
                issues.append("Initial without trailing period")

            # ---- inconsistent rank prefixes ----
            rank_prefixes = r"^(Gen\.|Genl|Lt-Gen|Maj-Gen|Brig-Gen|Col|Lt-Col|Maj|Capt|Cmdt|Asst-Cmdt-Gen|Cpl|Sgt|Pte)\s"
            rank_m = re.match(rank_prefixes, cmd)
            if rank_m:
                # Rank prefix present — check if surname, firstname follows
                after_rank = cmd[rank_m.end():]
                if "," not in after_rank and not re.match(r"^[A-Z][a-z]", after_rank):
                    issues.append("Rank prefix present but name format unclear")

        if not issues:
            fill_row(ws2, row_num,
                     [row_id, evt, cmd, "OK", ""],
                     fill=FILL_OK)
        else:
            fill = FILL_BAD if "Blank commander" in " ".join(issues) else FILL_WARN
            fill_row(ws2, row_num,
                     [row_id, evt, cmd, "; ".join(issues), suggestion],
                     fill=fill)
        row_num += 1

    # =========================================================================
    # SHEET 3 — Date Quality
    # =========================================================================
    print("Building Sheet 3: Date Quality …")
    ws3 = wb.create_sheet("3 Date Quality")

    headers3 = ["Row ID", "event_type", "date_start", "date_end", "Issue Type"]
    widths3  = [8, 16, 14, 14, 50]
    style_header_row(ws3, headers3, widths3)

    # Events that must be dated
    must_be_dated = {"engagement", "battle", "defeat", "capture", "skirmish", "execution"}

    # Tally date_start values for cluster detection
    date_start_counts = Counter(r.get("date_start", "").strip() for r in rows if r.get("date_start", "").strip())
    suspicious_dates  = {d for d, c in date_start_counts.items() if c > 10}

    row_num = 2
    for r in rows:
        row_id = r.get("id", "")
        evt    = r.get("event_type", "").strip().lower()
        ds_raw = r.get("date_start", "").strip()
        de_raw = r.get("date_end",   "").strip()

        issues = []

        # Format checks
        if ds_raw and not is_valid_date_fmt(ds_raw):
            issues.append(f"date_start not YYYY-MM-DD: '{ds_raw}'")
        if de_raw and not is_valid_date_fmt(de_raw):
            issues.append(f"date_end not YYYY-MM-DD: '{de_raw}'")

        # Blank date_start on engagements
        if not ds_raw and evt in must_be_dated:
            issues.append(f"Blank date_start on '{evt}' row — must be dated")

        # Impossible range (start > end)
        ds = parse_date(ds_raw)
        de = parse_date(de_raw)
        if ds and de and ds > de:
            issues.append(f"date_start ({ds_raw}) > date_end ({de_raw}) — impossible range")

        # Suspicious date cluster
        if ds_raw in suspicious_dates:
            issues.append(f"Suspicious cluster: {date_start_counts[ds_raw]} rows share date_start '{ds_raw}'")

        if not issues:
            fill_row(ws3, row_num,
                     [row_id, evt, ds_raw, de_raw, "OK"],
                     fill=FILL_OK)
        else:
            fill = FILL_BAD if any(
                kw in " ".join(issues) for kw in ["impossible", "Blank date_start on 'engagement'"]
            ) else FILL_WARN
            fill_row(ws3, row_num,
                     [row_id, evt, ds_raw, de_raw, "; ".join(issues)],
                     fill=fill)
        row_num += 1

    # =========================================================================
    # SHEET 4 — Place Name Consistency
    # =========================================================================
    print("Building Sheet 4: Place Name Consistency …")
    ws4 = wb.create_sheet("4 Place Names")

    headers4 = ["Row ID", "Field", "Value", "Issue Type", "Possible Matches"]
    widths4  = [8, 14, 36, 36, 40]
    style_header_row(ws4, headers4, widths4)

    # Broad geographic areas — not towns
    broad_places = re.compile(
        r"^(Orange Free State|Cape Colony|Transvaal|Natal|South Africa|"
        r"Orange River Colony|Rhodesia|Basutoland|Bechuanaland|Swaziland|"
        r"Griqualand|eastern Cape|western Cape|north-west Cape|Cape Midlands)$",
        re.IGNORECASE
    )

    # Build normalised lookup of all place values to detect near-duplicates
    all_places_raw = []
    for r in rows:
        for field in ("action_place", "from_place", "to_place"):
            val = r.get(field, "").strip()
            if val:
                all_places_raw.append(val)

    # Cluster near-duplicate place names (group by first 6 chars + fuzzy)
    place_counter = Counter(all_places_raw)
    unique_places = list(place_counter.keys())

    # Build near-dup clusters: for each place find others with score >=85
    near_dup_map = {}   # place -> list of near-dups
    for i, p1 in enumerate(unique_places):
        matches = []
        for j, p2 in enumerate(unique_places):
            if i == j:
                continue
            if abs(len(p1) - len(p2)) > 12:
                continue
            # Quick prefix check before full fuzzy
            if p1[:3].lower() == p2[:3].lower():
                score = fuzzy_ratio(p1, p2)
                if score >= 85 and score < 100:
                    matches.append(p2)
        if matches:
            near_dup_map[p1] = matches

    row_num = 2
    for r in rows:
        row_id     = r.get("id", "")
        from_place = r.get("from_place",  "").strip()
        to_place   = r.get("to_place",    "").strip()
        act_place  = r.get("action_place","").strip()
        evt        = r.get("event_type",  "").strip().lower()

        written = False

        # from == to on movement rows
        if evt in ("advance", "move", "march", "redeployment", "deployment", "rail-move", "retreat") \
                and from_place and to_place and from_place.lower() == to_place.lower():
            fill_row(ws4, row_num,
                     [row_id, "from/to", f"{from_place} → {to_place}",
                      "from_place == to_place on movement row", ""],
                     fill=FILL_BAD)
            row_num += 1
            written = True

        # action_place checks
        if act_place:
            issues = []
            # Broad province/country
            if broad_places.match(act_place):
                issues.append("Broad geographic area, not a specific town")

            # Near-duplicate spellings
            near_dups = near_dup_map.get(act_place, [])
            if near_dups:
                issues.append(f"Near-duplicate spelling found")

            # Case inconsistency (all lower when elsewhere capitalised)
            if act_place != act_place.title() and act_place == act_place.lower() and len(act_place) > 3:
                issues.append("All-lowercase place name")

            if issues:
                fill = FILL_BAD if "Broad" in " ".join(issues) else FILL_WARN
                fill_row(ws4, row_num,
                         [row_id, "action_place", act_place,
                          "; ".join(issues),
                          ", ".join(near_dups[:5]) if near_dups else ""],
                         fill=fill)
                row_num += 1
                written = True

        if not written:
            fill_row(ws4, row_num,
                     [row_id, "action_place", act_place or "(blank)", "OK", ""],
                     fill=FILL_OK)
            row_num += 1

    # =========================================================================
    # SHEET 5 — Source Quality
    # =========================================================================
    print("Building Sheet 5: Source Quality …")
    ws5 = wb.create_sheet("5 Source Quality")

    headers5 = ["Row ID", "source (raw)", "note (raw)", "Issue Type", "Source Tally"]
    widths5  = [8, 50, 30, 40, 20]
    style_header_row(ws5, headers5, widths5)

    # First pass: tally source tokens
    source_tokens = Counter()
    for r in rows:
        src = r.get("source", "").strip()
        for part in re.split(r"[;,]", src):
            part = part.strip()
            if part:
                source_tokens[part] += 1

    # Write tally as a header block first (rows 2 onward in col 5 as info)
    # We'll embed it in the data rows as a column note on first occurrence

    row_num = 2
    for r in rows:
        row_id = r.get("id", "")
        src    = r.get("source", "").strip()
        note   = r.get("note",   "").strip()

        issues = []

        # Very short source
        if src and len(src) < 5:
            issues.append(f"Source extremely short (<5 chars): '{src}'")

        # Bare angloboerwar.com with no chapter/page reference
        if re.search(r"angloboerwar", src, re.I):
            # Check if there's a chapter or page qualifier
            if not re.search(r"(ch\.|chapter|p\.|page|vol|jnl|pp\.|ref)", src, re.I):
                issues.append("angloboerwar.com cited without chapter/page/vol qualifier")

        # Source same as note
        if src and note and src.lower() == note.lower():
            issues.append("source == note (duplicate content)")

        # Blank source
        if not src:
            issues.append("No source at all")

        # Tally for this row's primary source
        primary = re.split(r"[;,]", src)[0].strip() if src else ""
        tally_str = f"×{source_tokens.get(primary, 0)}" if primary else ""

        if not issues:
            fill_row(ws5, row_num,
                     [row_id, src, note, "OK", tally_str],
                     fill=FILL_OK)
        else:
            fill = FILL_BAD if "No source" in " ".join(issues) else FILL_WARN
            fill_row(ws5, row_num,
                     [row_id, src, note, "; ".join(issues), tally_str],
                     fill=fill)
        row_num += 1

    # Append source frequency summary at the bottom
    row_num += 1
    ws5.cell(row=row_num, column=1, value="SOURCE FREQUENCY SUMMARY").font = FONT_BOLD
    row_num += 1
    ws5.cell(row=row_num, column=1, value="Source Token").font = FONT_BOLD
    ws5.cell(row=row_num, column=2, value="Count").font = FONT_BOLD
    row_num += 1
    for tok, cnt in source_tokens.most_common(40):
        ws5.cell(row=row_num, column=1, value=tok)
        ws5.cell(row=row_num, column=2, value=cnt)
        row_num += 1

    # =========================================================================
    # SHEET 6 — Description Quality
    # =========================================================================
    print("Building Sheet 6: Description Quality …")
    ws6 = wb.create_sheet("6 Description Quality")

    headers6 = ["Row ID", "date_start", "Description (first 80 chars)", "Issue Type", "Detail"]
    widths6  = [8, 12, 55, 40, 40]
    style_header_row(ws6, headers6, widths6)

    PLACEHOLDER_RE = re.compile(
        r"\b(TBD|TODO|placeholder|n/a|N/A|unknown|UNKNOWN|tbd|todo)\b"
    )
    # Build description index for duplicate / template detection
    desc_exact  = Counter(r.get("description", "").strip() for r in rows
                          if r.get("description", "").strip())
    desc_prefix = Counter(r.get("description", "").strip()[:30] for r in rows
                          if r.get("description", "").strip())

    # Year-in-description regex
    year_re = re.compile(r"\bin (1[89]\d{2})\b")

    row_num = 2
    for r in rows:
        row_id  = r.get("id", "")
        desc    = r.get("description", "").strip()
        ds_raw  = r.get("date_start", "").strip()
        ds      = parse_date(ds_raw)

        issues  = []
        detail  = ""

        if not desc:
            issues.append("Blank description")
        else:
            # Exact duplicate across rows
            if desc_exact[desc] > 1:
                issues.append(f"Exact duplicate description (×{desc_exact[desc]} rows)")

            # Template repetition (same 30-char prefix used >3 times)
            prefix = desc[:30]
            if desc_prefix[prefix] > 3 and len(prefix) == 30:
                issues.append(f"Template prefix repeated ×{desc_prefix[prefix]}")

            # Placeholder phrases
            m = PLACEHOLDER_RE.search(desc)
            if m:
                issues.append(f"Placeholder phrase: '{m.group()}'")

            # Wrong year mentioned in description vs date_start
            if ds:
                year_start = ds.year
                for ym in year_re.finditer(desc):
                    mentioned = int(ym.group(1))
                    # Allow ±1 year slop for campaigns that span year-ends
                    if abs(mentioned - year_start) > 1:
                        issues.append(f"Description mentions year {mentioned} but date_start is {year_start}")
                        detail = f"Desc year: {mentioned}, date_start year: {year_start}"

        desc_preview = desc[:80] if desc else "(blank)"

        if not issues:
            fill_row(ws6, row_num,
                     [row_id, ds_raw, desc_preview, "OK", ""],
                     fill=FILL_OK)
        else:
            fill = FILL_BAD if "Blank" in " ".join(issues) or "Placeholder" in " ".join(issues) \
                   else FILL_WARN
            fill_row(ws6, row_num,
                     [row_id, ds_raw, desc_preview, "; ".join(issues), detail],
                     fill=fill)
        row_num += 1

    # =========================================================================
    # SHEET 7 — QSA Clasp Coverage
    # =========================================================================
    print("Building Sheet 7: QSA Clasp Coverage …")
    ws7 = wb.create_sheet("7 QSA Clasp Coverage")

    headers7 = ["QSA Clasp", "Status", "Matching Row IDs", "Matching action_place / description snippets"]
    widths7  = [30, 14, 20, 60]
    style_header_row(ws7, headers7, widths7)

    # Build a searchable corpus per row: action_place + description + event_type
    def row_corpus(r):
        return " ".join([
            r.get("action_place", ""),
            r.get("description", ""),
            r.get("event_type", ""),
            r.get("from_place", ""),
            r.get("to_place", ""),
            r.get("note", ""),
        ]).upper()

    # Precompute corpora
    corpora = [(r.get("id",""), row_corpus(r), r) for r in rows]

    # Keyword synonyms to help find coverage of each clasp
    CLASP_KEYWORDS = {
        "CAPE COLONY":            ["CAPE COLONY", "CAPE MIDLAND", "CAPE REBEL"],
        "NATAL":                  ["NATAL", "LADYSMITH", "TUGELA", "DUNDEE", "ESTCOURT", "NEWCASTLE"],
        "RHODESIA":               ["RHODESIA", "MAFEKING"],
        "DEFENCE OF KIMBERLEY":   ["KIMBERLEY", "KIMBERLEY DEFEND", "SIEGE OF KIMBERLEY"],
        "TALANA":                 ["TALANA"],
        "ELANDSLAAGTE":           ["ELANDSLAAGTE", "ELANDS LAAGTE"],
        "DEFENCE OF LADYSMITH":   ["DEFENCE OF LADYSMITH", "SIEGE OF LADYSMITH", "LADYSMITH GARRISON"],
        "BELMONT":                ["BELMONT"],
        "MODDER RIVER":           ["MODDER RIVER", "MODDER"],
        "TUGELA HEIGHTS":         ["TUGELA HEIGHT", "PIETERS HILL", "TUGELA"],
        "RELIEF OF KIMBERLEY":    ["RELIEF OF KIMBERLEY", "KIMBERLEY RELIEF", "KIMBERLEY"],
        "PAARDEBERG":             ["PAARDEBERG"],
        "RELIEF OF LADYSMITH":    ["RELIEF OF LADYSMITH", "RELIEVE LADYSMITH"],
        "DRIEFONTEIN":            ["DRIEFONTEIN"],
        "WEPENER":                ["WEPENER"],
        "DEFENCE OF MAFEKING":    ["MAFEKING", "DEFENCE OF MAFEKING"],
        "JOHANNESBURG":           ["JOHANNESBURG"],
        "DIAMOND HILL":           ["DIAMOND HILL", "DONKERHOEK"],
        "WITTEBERGEN":            ["WITTEBERGEN", "BRANDWATER BASIN"],
        "BELFAST":                ["BELFAST", "BERGENDAL"],
        "TRANSVAAL":              ["TRANSVAAL", "PRETORIA", "JOHANNESBURG", "BELFAST"],
        "ORANGE FREE STATE":      ["ORANGE FREE STATE", "BLOEMFONTEIN", "OFS"],
        "LAING'S NEK":            ["LAING", "LAINGS NEK", "ALLEMAN"],
        "RELIEF OF MAFEKING":     ["RELIEF OF MAFEKING", "MAFEKING RELIEF"],
        "SOUTH AFRICA 1901":      ["1901"],
        "SOUTH AFRICA 1902":      ["1902"],
    }

    row_num = 2
    for clasp in QSA_CLASPS:
        keywords = CLASP_KEYWORDS.get(clasp, [clasp])
        matching_ids  = []
        matching_snip = []

        for row_id, corpus, r in corpora:
            hit = any(kw in corpus for kw in keywords)
            if hit:
                matching_ids.append(str(row_id))
                act = r.get("action_place", "").strip()
                if act:
                    matching_snip.append(act[:40])

        if matching_ids:
            status = "COVERED"
            fill   = FILL_OK
        else:
            status = "NOT COVERED"
            fill   = FILL_BAD

        fill_row(ws7, row_num,
                 [clasp,
                  status,
                  ", ".join(matching_ids[:15]) + ("…" if len(matching_ids) > 15 else ""),
                  "; ".join(dict.fromkeys(matching_snip[:8]))],
                 fill=fill)
        row_num += 1

    # Annotation below the clasp table
    row_num += 1
    note_cell = ws7.cell(row=row_num, column=1,
                         value="NOTE: 'NOT COVERED' means no CSV row's action_place/description/event matches "
                               "the clasp keywords. Add explicit engagement rows or note the gap in the data.")
    note_cell.font = Font(name="Calibri", italic=True, size=9, color="444444")
    ws7.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)

    # =========================================================================
    # SAVE
    # =========================================================================
    wb.save(OUT_PATH)
    print(f"\nDone → {OUT_PATH}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    # Validate paths before running
    for p in (CSV_PATH, JSON_PATH):
        if not os.path.exists(p):
            sys.exit(f"ERROR: expected file not found: {p!r}\n"
                     f"Run this script from the boer-war-eastern-cape directory.")
    main()
