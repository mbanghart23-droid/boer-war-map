"""
Movement gap audit — finds chronological jumps (>30 days) between events for each unit,
flags unknown bridging movements, and notes what we can/can't fill.
"""
import json, csv
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = __import__('pathlib').Path(__file__).parent.parent

gj = json.load(open(HERE/"docs"/"data"/"events.geojson", encoding="utf-8"))

# ── build per-regiment timeline ───────────────────────────────────────────────
unit_events = defaultdict(list)
for f in gj["features"]:
    p = f["properties"]
    if p["kind"] != "event":
        continue
    for g in (p.get("groups") or []):
        unit_events[g[0]].append(p)

# ── count event coverage in the GeoJSON (for context) ────────────────────────
rows_csv = list(csv.DictReader(open(HERE/"data"/"movements.csv", encoding="utf-8")))

def parse_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], "%Y-%m-%d")
    except: return None

# ── find gaps ─────────────────────────────────────────────────────────────────
GAP_THRESHOLD = 30   # days — gaps below this are noise

all_gaps = []
for unit in sorted(unit_events.keys()):
    evts = sorted(unit_events[unit], key=lambda e: e.get("date_start",""))
    for i in range(len(evts)-1):
        a, b = evts[i], evts[i+1]
        d1 = parse_date(a.get("date_end") or a.get("date_start",""))
        d2 = parse_date(b.get("date_start",""))
        if not d1 or not d2: continue
        days = (d2 - d1).days
        if days < GAP_THRESHOLD: continue
        if a.get("place") == b.get("place") and days < 90: continue  # same place, short = ok
        all_gaps.append({
            "unit": unit,
            "side": a.get("side",""),
            "gap_days": days,
            "from_date": a.get("date_end") or a.get("date_start",""),
            "from_place": a.get("place",""),
            "from_force": a.get("force",""),
            "from_event": a.get("event_type",""),
            "from_id": a.get("id",""),
            "to_date": b.get("date_start",""),
            "to_place": b.get("place",""),
            "to_force": b.get("force",""),
            "to_event": b.get("event_type",""),
            "to_id": b.get("id",""),
            "priority": "HIGH" if days > 180 else ("MEDIUM" if days > 90 else "LOW"),
            "notes": "",
        })

# ── specific notes for known gaps ────────────────────────────────────────────
KNOWN_NOTES = {
    ("17th Lancers", "1900-04-30", "1900-11-01"): "De Aar → Bloemfontein leg missing; they joined Roberts via Modder River. Then Diamond Hill → SE ORC under Herbert is a ~5 month gap — regiment likely doing column work in Transvaal/ORC not yet documented",
    ("17th Lancers", "1900-06-11", "1900-11-01"): "After Diamond Hill regiment moved SE to pursue De Wet toward Cape; Herbert's column assembled Bethulie-Springfontein area Nov-Dec 1900. Route: Pretoria → Bloemfontein → Bethulie rail line (not yet mapped)",
    ("17th Lancers", "1901-06-07", "1901-08-13"): "After Wyndham's night attack at Ruigtevlei (Jul 1901) regiment redistriicted under Gorringe — short gap, likely local column repositioning",
    ("Kritzinger commando", "1901-06-20", "1901-12-11"): "Kritzinger arrested Jun 1901 by British at Graaff-Reinet; Nieuwoudt took command. Commando scattered & regrouped near Colesberg by Dec 1901 — gap is political/command transition",
    ("Smuts commando", "1901-09-17", "1902-04-01"): "After Elands River Poort (Sep 1901) Smuts's 300 men raided deep west — Calvinia, Clanwilliam, O'kiep. Most of this period is on the map as Hertzog raid (separate) and Smuts Namaqualand (row 701/702). Check coverage.",
    ("Cape Mounted Rifles", "1900-04-09", "1901-09-05"): "498-day gap is very large — CMR served all over ORC and Transvaal after Wepener relief. Major documentation hole; Groenkloof (Sep 1901) is the Elands River Poort ambush site",
    ("Brabant's Horse", "1900-04-09", "1901-01-11"): "After Wepener siege Brabant's Horse served in ORC columns pursuing De Wet — no EC events for 8+ months. Needs research: were any squadrons in the EC during this period?",
}

for g in all_gaps:
    key1 = (g["unit"], g["from_date"][:10] if g["from_date"] else "", g["to_date"][:10] if g["to_date"] else "")
    key2 = (g["unit"], (g["from_date"] or "")[:7], (g["to_date"] or "")[:7])
    for k, note in KNOWN_NOTES.items():
        if k[0] == g["unit"] and k[1][:7] == (g["from_date"] or "")[:7]:
            g["notes"] = note

# ── write Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: All gaps ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "All Gaps"

HEADERS = ["Unit","Side","Gap Days","Priority","From Date","From Place","From Force","From Type","To Date","To Place","To Force","To Type","Notes","From ID","To ID"]
ws.append(HEADERS)

H_FILL = PatternFill("solid", fgColor="1e3a5f")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(1, col)
    c.fill = H_FILL; c.font = H_FONT; c.alignment = Alignment(wrap_text=True)

PRIORITY_COLORS = {"HIGH":"FF4444","MEDIUM":"FF8800","LOW":"FFCC44"}
for g in sorted(all_gaps, key=lambda x: (-x["gap_days"], x["unit"])):
    row = [g["unit"],g["side"],g["gap_days"],g["priority"],
           g["from_date"],g["from_place"],g["from_force"],g["from_event"],
           g["to_date"],g["to_place"],g["to_force"],g["to_event"],
           g["notes"],g["from_id"],g["to_id"]]
    ws.append(row)
    r = ws.max_row
    color = PRIORITY_COLORS.get(g["priority"],"FFFFFF")
    for col in range(1, 5):
        ws.cell(r, col).fill = PatternFill("solid", fgColor=color+"22")
    ws.cell(r, 3).alignment = Alignment(horizontal="center")
    ws.cell(r, 4).font = Font(bold=True, color=PRIORITY_COLORS.get(g["priority"],"000000"))
    if g["notes"]:
        ws.cell(r, 13).font = Font(italic=True, color="1a4ab9")

COL_WIDTHS = [28,8,10,9,12,22,32,12,12,22,32,12,55,8,8]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Sheet 2: 17th Lancers detail ─────────────────────────────────────────────
ws2 = wb.create_sheet("17th Lancers Journey")
ws2.append(["ID","Date Start","Date End","Place","Force / Column","Event Type","Description","Source","Gap to next (days)","Gap notes"])
h2f = PatternFill("solid", fgColor="2d1a00"); h2fn = Font(bold=True, color="FFD700", size=10)
for col in range(1, 11):
    ws2.cell(1, col).fill = h2f; ws2.cell(1, col).font = h2fn

lancer_rows = list(csv.DictReader(open(HERE/"data"/"movements.csv", encoding="utf-8")))
lancer_events = [r for r in lancer_rows if "17th" in r.get("units","") or "17th Lancers" in r.get("force","")]
lancer_events.sort(key=lambda r: r["date_start"])

GAP_NOTES_17 = {
    "1900-04-30": "MISSING: De Aar → Bloemfontein (via Modder River). Regiment joined Roberts's northern advance but this leg is not documented. Likely rail or march Modder River → Bloemfontein ~May 1900.",
    "1900-06-11": "MISSING: Diamond Hill → SE ORC (Bethulie area). 142-day gap. After Donkerhoek the regiment operated in the Transvaal/ORC but exact column assignments unknown. Herbert assembled his column ~Oct-Nov 1900 at Bethulie.",
    "1901-05-31": "MISSING: ORC → Eastern Cape. Regiment moved from ORC back into the Cape Colony (entering around Norvalspont) for the Gorringe/Wyndham column phase. Exact route unknown.",
}

for r in lancer_events:
    d_end = r.get("date_end") or r.get("date_start")
    gap_note = GAP_NOTES_17.get(d_end[:10] if d_end else "", "")
    ws2.append([r["id"], r["date_start"], r["date_end"], r["action_place"] or r["from_place"],
                r["force"], r["event_type"], r["description"][:120], r["source"], "", gap_note])
    rn = ws2.max_row
    if gap_note:
        ws2.cell(rn, 9).value = "???"
        ws2.cell(rn, 10).font = Font(bold=True, color="CC0000")
        ws2.cell(rn, 10).fill = PatternFill("solid", fgColor="FFF0F0")

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 12; ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 22; ws2.column_dimensions["E"].width = 32
ws2.column_dimensions["F"].width = 14; ws2.column_dimensions["G"].width = 50
ws2.column_dimensions["H"].width = 20; ws2.column_dimensions["I"].width = 8
ws2.column_dimensions["J"].width = 60
ws2.freeze_panes = "A2"

# ── Sheet 3: High priority gaps to research ───────────────────────────────────
ws3 = wb.create_sheet("Research List")
ws3.append(["Unit","Gap Days","From","To","What we need to find","Source suggestions"])
h3f = PatternFill("solid", fgColor="0f3d1a"); h3fn = Font(bold=True, color="90EE90", size=10)
for col in range(1,7):
    ws3.cell(1,col).fill = h3f; ws3.cell(1,col).font = h3fn

RESEARCH = [
    ["17th Lancers", 142, "Diamond Hill Jun 1900", "Bethulie Nov 1900",
     "Exact column assignments after Diamond Hill. How did they get from Pretoria area to Bethulie (SE ORC)?",
     "Regimental history '17th (Duke of Cambridge's Own) Lancers'; Haig diary (Haig Papers, NLS); Conan Doyle chs XVII-XIX"],
    ["17th Lancers", "~30", "De Aar Apr 1900", "Bloemfontein May 1900",
     "Missing link: De Aar → Roberts's advance. Did they go via Modder River or direct rail?",
     "Conan Doyle ch XIV; Roberts's dispatch Apr 1900; regimental diary"],
    ["Cape Mounted Rifles", 498, "Wepener Apr 1900", "Groenkloof Sep 1901",
     "CMR served widely in ORC/Transvaal. Any EC service 1900-1901?",
     "CMR regimental history; Warwick 'Black People and the South African War'"],
    ["Brabant's Horse", 261, "Wepener Apr 1900", "Murraysburg Jan 1901",
     "Where were Brabant's squadrons Jun-Dec 1900? Known to have served in ORC.",
     "Benyon 'Proconsul and Paramountcy'; Pakenham 'Boer War'"],
    ["Kritzinger commando", 173, "Graaff-Reinet Jun 1901", "Colesberg Dec 1901",
     "After Kritzinger's capture (Jun 1901) who commanded? Where did they regroup?",
     "Nasson 'Abraham Esau's War'; Grundlingh 'The Dynamics of Treason'"],
    ["9th Lancers", 486, "Paardeberg Feb 1900", "Richmond Jun 1901",
     "486-day gap — where were 9th Lancers 1900-1901? Known to have served in Transvaal.",
     "Regimental history '9th (Queen's Royal) Lancers'"],
    ["Smuts commando", 196, "Elands River Poort Sep 1901", "Okiep Apr 1902",
     "Smuts's western Cape raid Sep 1901-May 1902 is only partially mapped. Calvinia, Clanwilliam, Springbok legs missing.",
     "Hancock 'Smuts Vol 1'; Kruger 'Goodbye Dolly Gray' ch 30; Smuts 'Memoirs of the Boer War'"],
    ["Cape Police", 345, "Burgersdorp Mar 1900", "Cradock Feb 1901",
     "Cape Police served in multiple columns. Where were they Mar 1900 - Feb 1901?",
     "Rauch 'The History of the Cape Police'; SAMAD Cape Police records"],
    ["South African Light Horse", 375, "Ladysmith Jan 1900", "Murraysburg Jan 1901",
     "SALH served in Natal, then in EC guerrilla columns. Missing the transition.",
     "Conan Doyle; Amery 'The Times History of the War in South Africa' Vol 4"],
]

for row in RESEARCH:
    ws3.append(row)
    rn = ws3.max_row
    ws3.cell(rn, 5).alignment = Alignment(wrap_text=True)
    ws3.cell(rn, 6).alignment = Alignment(wrap_text=True)
    ws3.cell(rn, 6).font = Font(italic=True, color="1a4ab9")

ws3.column_dimensions["A"].width = 28; ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 20; ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 55; ws3.column_dimensions["F"].width = 55
ws3.freeze_panes = "A2"

out = HERE/"tools"/"boer_war_gap_audit.xlsx"
wb.save(out)
print(f"Saved {out}")
print(f"Total gaps found: {len(all_gaps)} ({sum(1 for g in all_gaps if g['priority']=='HIGH')} HIGH)")
