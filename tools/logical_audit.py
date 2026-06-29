"""
Deep logical audit of movements.csv — identifies historical/geographic impossibilities,
contradictions, and suspicious data for researcher review.

Checks:
  A. Geographic impossibility  — travel speed too fast given distance
  B. Battle truth table        — known dates/places for 30 major engagements
  C. Text vs event_type        — description implies different outcome than event_type
  D. Side vs force name        — Boer unit name on British side etc.
  E. Unit arrival dates        — events before regiment known to be in SA
  F. Commander assignment      — commander known to be elsewhere / wrong unit
  G. Chronological paradox     — date_end before date_start, or events that can't follow each other
  H. Theater mismatch          — Eastern Cape unit suddenly in Natal with no movement row
  I. Description contradictions— "victory" / "defeat" keywords vs event_type
  J. Known wrong coordinates   — famous battles placed far from historical location
"""
import csv, json, re, math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent.parent

rows   = list(csv.DictReader(open(HERE/"data"/"movements.csv", encoding="utf-8")))
gj     = json.load(open(HERE/"docs"/"data"/"events.geojson", encoding="utf-8"))

row_by_id  = {r["id"]: r for r in rows}
feat_by_id = {f["properties"]["id"]: f for f in gj["features"] if f["properties"]["kind"]=="event"}

issues = []

def add(check, severity, row_ids, field, description, historical_note=""):
    if isinstance(row_ids, str): row_ids = [row_ids]
    force = " / ".join(row_by_id.get(i,{}).get("force","")[:35] for i in row_ids[:2])
    issues.append({
        "check": check, "severity": severity,
        "ids": ", ".join(row_ids[:3]),
        "force": force,
        "field": field,
        "description": description,
        "historical_note": historical_note,
    })

def parse_date(s):
    if not s: return None
    try: return datetime.strptime(s[:10], "%Y-%m-%d")
    except: return None

def haversine(c1, c2):
    """Distance in km between two [lon,lat] coords."""
    lon1,lat1 = math.radians(c1[0]),math.radians(c1[1])
    lon2,lat2 = math.radians(c2[0]),math.radians(c2[1])
    dlat = lat2-lat1; dlon = lon2-lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 6371 * 2 * math.asin(math.sqrt(a))

# ═══════════════════════════════════════════════════════════════════════════════
# A. GEOGRAPHIC TRAVEL SPEED
# Max realistic speeds:
#   Horse/cavalry column: ~40 km/day sustained, 70 km/day sprint
#   Infantry march:       ~25 km/day sustained
#   Railway:              ~300 km/day
#   Boer commando:        ~50 km/day mounted
# We flag anything > 120 km/day as suspicious (allows for mixed rail+march)
# ═══════════════════════════════════════════════════════════════════════════════
MAX_KM_PER_DAY = 120

# Build per-force timeline with coordinates
force_timeline = defaultdict(list)
for fid, feat in feat_by_id.items():
    p = feat["properties"]
    coords = feat["geometry"]["coordinates"]
    d = parse_date(p.get("date_start",""))
    if d:
        force_timeline[p.get("force","")].append({
            "id": fid, "date": d, "date_end": parse_date(p.get("date_end","")),
            "coords": coords, "place": p.get("place",""),
            "event_type": p.get("event_type",""),
        })

for force, timeline in force_timeline.items():
    timeline.sort(key=lambda e: e["date"])
    for i in range(len(timeline)-1):
        a, b = timeline[i], timeline[i+1]
        d1 = a.get("date_end") or a["date"]
        d2 = b["date"]
        days = max((d2 - d1).days, 1)
        dist = haversine(a["coords"], b["coords"])
        kpd = dist / days
        if kpd > MAX_KM_PER_DAY and dist > 50:
            add("A. Travel Speed", "HIGH",
                [a["id"], b["id"]], "coords/date",
                f"'{force}': {dist:.0f}km in {days}d = {kpd:.0f}km/day "
                f"(max realistic ~120). "
                f"{a['place']} [{a['date'].strftime('%Y-%m-%d')}] → "
                f"{b['place']} [{b['date'].strftime('%Y-%m-%d')}]",
                "Either one of the dates is wrong, the place is wrong, or a rail movement is undocumented between these two events.")

# ═══════════════════════════════════════════════════════════════════════════════
# B. BATTLE TRUTH TABLE — known correct date + location for major engagements
# Format: (name, correct_date_start, correct_date_end, correct_place, lat, lon, radius_km)
# radius_km = how close the map marker should be to the real location
# ═══════════════════════════════════════════════════════════════════════════════
TRUTH = [
    # (display_name, date_start, date_end, place_name, lat, lon, radius_km)
    ("Talana Hill",         "1899-10-20","1899-10-20","Dundee",      -28.16, 30.22, 15),
    ("Elandslaagte",        "1899-10-21","1899-10-21","Elandslaagte",-28.32, 30.14, 10),
    ("Ladysmith siege",     "1899-11-02","1900-02-28","Ladysmith",   -28.55, 29.77, 20),
    ("Belmont",             "1899-11-23","1899-11-23","Belmont",     -29.32, 24.35, 20),
    ("Graspan",             "1899-11-25","1899-11-25","Graspan",     -29.14, 24.43, 20),
    ("Modder River",        "1899-11-28","1899-11-28","Modder River",-29.03, 24.62, 15),
    ("Stormberg",           "1899-12-10","1899-12-10","Stormberg",   -31.26, 26.57, 15),
    ("Magersfontein",       "1899-12-11","1899-12-11","Magersfontein",-29.03,24.88, 15),
    ("Colenso",             "1899-12-15","1899-12-15","Colenso",     -28.73, 29.83, 15),
    ("Spion Kop",           "1900-01-24","1900-01-24","Spion Kop",   -28.57, 29.53, 10),
    ("Paardeberg",          "1900-02-18","1900-02-27","Paardeberg",  -29.05, 25.00, 15),
    ("Kimberley relief",    "1900-02-15","1900-02-15","Kimberley",   -28.73, 24.77, 15),
    ("Ladysmith relief",    "1900-02-28","1900-02-28","Ladysmith",   -28.55, 29.77, 20),
    ("Bloemfontein falls",  "1900-03-13","1900-03-13","Bloemfontein",-29.12, 26.21, 15),
    ("Sannaspos",           "1900-03-31","1900-03-31","Sannaspos",   -29.10, 26.27, 20),
    ("Wepener siege",       "1900-04-09","1900-04-25","Wepener",     -29.72, 27.03, 15),
    ("Mafeking relief",     "1900-05-17","1900-05-17","Mafeking",    -25.85, 25.65, 20),
    ("Pretoria falls",      "1900-06-05","1900-06-05","Pretoria",    -25.75, 28.23, 20),
    ("Diamond Hill",        "1900-06-11","1900-06-12","Diamond Hill",-25.79, 28.46, 20),
    ("Bergendal",           "1900-08-27","1900-08-27","Bergendal",   -25.78, 30.24, 15),
    ("Nooitgedacht",        "1900-12-13","1900-12-13","Nooitgedacht",-25.73, 27.49, 20),
    ("Groenkop",            "1901-12-25","1901-12-25","Groenkop",    -27.60, 29.08, 25),
    ("Tweebosch",           "1902-03-07","1902-03-07","Tweebosch",   -26.81, 23.63, 25),
    ("Rooiwal",             "1902-04-11","1902-04-11","Rooiwal",     -25.50, 27.38, 25),
    ("Elands River Poort",  "1901-09-17","1901-09-17","Elands River Poort",-32.32,24.97,20),
    ("Bakenlaagte",         "1901-10-30","1901-10-30","Bakenlaagte", -26.18, 30.05, 20),
    ("Blood River Poort",   "1901-09-17","1901-09-17","Blood River Poort",-27.20,30.00,25),
    ("Itala",               "1901-09-26","1901-09-26","Itala",       -27.38, 30.90, 20),
]

TRUTH_LOOKUP = {}
for t in TRUTH:
    TRUTH_LOOKUP[t[0].lower()] = t

# Check rows against truth table by matching force/description keywords
for r in rows:
    force_lower  = r.get("force","").lower()
    desc_lower   = r.get("description","").lower()
    combined     = force_lower + " " + desc_lower

    for name, t_ds, t_de, t_place, t_lat, t_lon, radius in TRUTH:
        keywords = name.lower().replace("'","").split()
        if not all(kw in combined for kw in keywords if len(kw)>3):
            continue

        # Date check
        ds = parse_date(r.get("date_start",""))
        td = parse_date(t_ds)
        if ds and td and abs((ds-td).days) > 3:
            add("B. Wrong Date", "HIGH", [r["id"]], "date_start",
                f"'{r['force']}' matches '{name}' but date is {r['date_start']} — expected ~{t_ds}",
                f"Historical date: {t_ds}" + (f"–{t_de}" if t_de != t_ds else ""))

        # Location check (if on map)
        feat = feat_by_id.get(r["id"])
        if feat:
            coords = feat["geometry"]["coordinates"]
            dist = haversine(coords, [t_lon, t_lat])
            if dist > radius:
                add("B. Wrong Location", "HIGH", [r["id"]], "place/coords",
                    f"'{r['force']}' matches '{name}' but marker is {dist:.0f}km from historical location "
                    f"(should be near {t_place} [{t_lat:.2f},{t_lon:.2f}], "
                    f"marker is at [{coords[1]:.2f},{coords[0]:.2f}])",
                    f"Historical location: {t_place}. Check OVERRIDES/TOWNS coord for '{r.get('action_place','')}'.")

# ═══════════════════════════════════════════════════════════════════════════════
# C. TEXT VS EVENT_TYPE CONTRADICTIONS
# ═══════════════════════════════════════════════════════════════════════════════
DEFEAT_WORDS = re.compile(r"\brepulsed\b|\bdriven off\b|\bwithdrew\b|\bwithdrawn\b|\bretreat\b|\bforced back\b|\bcasualties\b.*\bbrit\b|\bblack week\b|\bdisaster\b|\bdebacle\b|\bfled\b", re.I)
VICTORY_WORDS= re.compile(r"\bdrove off\b|\bdefeated\b|\bcaptured\b|\bsurrendered\b|\brelieved\b|\bfell\b.*\bboer\b|\bboer.*\bfled\b", re.I)
ADVANCE_TYPES= {"advance","pursuit","drive"}
DEFEAT_TYPES = {"defeat","retreat","surrender"}

for r in rows:
    desc = r.get("description","")
    et   = r.get("event_type","")
    side = r.get("side","")

    if side=="British" and DEFEAT_WORDS.search(desc) and et in ADVANCE_TYPES:
        add("C. Text/Type Mismatch", "MEDIUM", [r["id"]], "event_type",
            f"Description implies British setback/defeat but event_type='{et}': \"{desc[:100]}\"",
            "Consider event_type='defeat' or 'retreat'")

    if side=="Boer" and VICTORY_WORDS.search(desc) and et in DEFEAT_TYPES:
        add("C. Text/Type Mismatch", "MEDIUM", [r["id"]], "event_type",
            f"Description implies Boer victory but event_type='{et}': \"{desc[:100]}\"",
            "Consider event_type='raid' or 'engagement'")

    # deployment rows with engagement descriptions
    if et == "deployment" and re.search(r"\battack\b|\bfight\b|\bengag\b|\bbattle\b|\bkilled\b|\bwounded\b", desc, re.I):
        add("C. Text/Type Mismatch", "MEDIUM", [r["id"]], "event_type",
            f"event_type='deployment' but description sounds like an engagement: \"{desc[:100]}\"",
            "Check whether this should be event_type='engagement' or 'battle'")

# ═══════════════════════════════════════════════════════════════════════════════
# D. SIDE vs FORCE NAME contradictions
# ═══════════════════════════════════════════════════════════════════════════════
BRITISH_MARKERS = re.compile(r"\bcolumn\b|\brigade\b|\bregiment\b|\blancers\b|\bhussars\b|\bdragoon\b|\byeomanry\b|\brifles\b|\bfusiliers\b|\bguards\b|\broyal\b|\bimperial\b|\bbattery\b|\bdivision\b|\bbattalion\b|\binfantry\b", re.I)
BOER_MARKERS    = re.compile(r"\bcommando\b|\bcommandant\b|\bveldt\b|\bveld\b|\bburghers?\b|\borange free state\b|\btransvaal\b", re.I)

for r in rows:
    force = r.get("force","")
    side  = r.get("side","")
    if side=="Boer" and BRITISH_MARKERS.search(force) and not BOER_MARKERS.search(force):
        add("D. Side/Force Mismatch", "MEDIUM", [r["id"]], "side",
            f"side='Boer' but force name sounds British: '{force}'",
            "Check whether side field is correct")
    if side=="British" and BOER_MARKERS.search(force):
        add("D. Side/Force Mismatch", "MEDIUM", [r["id"]], "side",
            f"side='British' but force name sounds Boer: '{force}'",
            "Check whether side field is correct")

# ═══════════════════════════════════════════════════════════════════════════════
# E. UNIT ARRIVAL DATES — units that appear before they could have arrived in SA
# Known arrival dates (earliest possible date in theatre)
# ═══════════════════════════════════════════════════════════════════════════════
ARRIVAL_DATES = {
    # regiment_pattern: (earliest_possible_date, note)
    r"17th lancer":         ("1900-03-10", "17th Lancers arrived Cape Town ~10 Mar 1900 on SS Victorian"),
    r"lovat.*scout":        ("1900-01-01", "Lovat's Scouts raised Dec 1899, arrived SA early 1900"),
    r"tasmanian":           ("1900-05-01", "Tasmanian Imperial Bushmen arrived SA ~May 1900"),
    r"new zealand":         ("1900-02-01", "NZ Mounted Rifles arrived Feb 1900"),
    r"9th lancer":          ("1899-10-01", "9th Lancers arrived with initial force Sep-Oct 1899"),
    r"gordon highlander":   ("1899-10-01", "Gordon Highlanders arrived Oct 1899 with Elandslaagte force"),
    r"imperial light horse":("1899-09-01", "ILH raised Natal Sep 1899"),
    r"imperial yeomanry":   ("1900-03-01", "First IY contingent arrived SA ~Mar 1900"),
    r"bethune.*mi|bethune.*mounted": ("1900-01-01", "Bethune's MI raised Natal Jan 1900"),
    r"nsw lancer|new south wales": ("1900-03-01", "NSW Lancers arrived Mar 1900"),
    r"scottish horse":      ("1900-03-01", "Scottish Horse raised and arrived Mar 1900"),
    r"kitchener.*scout|kitchener.*fight": ("1901-01-01", "Kitchener's Fighting Scouts formed Jan 1901"),
    r"damant.*horse":       ("1900-06-01", "Damant's Horse raised mid-1900"),
}

for r in rows:
    force_lower = r.get("force","").lower() + " " + r.get("units","").lower()
    ds = parse_date(r.get("date_start",""))
    if not ds: continue
    for pat, (earliest, note) in ARRIVAL_DATES.items():
        if re.search(pat, force_lower, re.I):
            ea = parse_date(earliest)
            if ea and ds < ea - timedelta(days=7):
                add("E. Pre-arrival Event", "HIGH", [r["id"]], "date_start",
                    f"'{r['force']}' dated {r['date_start']} but unit couldn't arrive before {earliest}",
                    note)

# ═══════════════════════════════════════════════════════════════════════════════
# F. KNOWN COMMANDER ASSIGNMENTS — spot-check obvious errors
# ═══════════════════════════════════════════════════════════════════════════════
CMD_TRUTH = [
    # (row_id_or_force_pattern, wrong_commander_pattern, correct_info)
    # Stormberg: Gatacre commanded, NOT Clements
    (r"stormberg|stormberg", r"clements", "Stormberg (10 Dec 1899) was commanded by Gatacre, not Clements"),
    # Magersfontein: Methuen (overall), Highland Brigade under Wauchope (killed)
    (r"magersfontein", r"french|buller|roberts", "Magersfontein was Methuen's action; Highland Brigade under Wauchope (killed)"),
    # Colenso: Buller
    (r"colenso", r"methuen|french|gatacre", "Colenso was Buller's action"),
    # Spion Kop: Buller overall, Warren led the assault, Thorneycroft on the summit
    (r"spion kop", r"methuen|french|gatacre|roberts", "Spion Kop: Buller/Warren commanded"),
    # Paardeberg: Roberts/Kitchener
    (r"paardeberg", r"buller|methuen|french", "Paardeberg: Roberts ordered it, Kitchener directed the assault"),
    # Sannaspos: De Wet ambushed Broadwood
    (r"sannaspos|koorn spruit|koornspruit", r"botha|joubert|de la rey", "Sannaspos was De Wet's operation (with Eloff)"),
]

for r in rows:
    combined = (r.get("force","") + " " + r.get("description","") + " " + r.get("action_place","")).lower()
    cmd_lower = r.get("commander","").lower()
    for place_pat, wrong_cmd_pat, correct in CMD_TRUTH:
        if re.search(place_pat, combined) and re.search(wrong_cmd_pat, cmd_lower):
            add("F. Wrong Commander", "HIGH", [r["id"]], "commander",
                f"'{r['commander']}' appears for '{r['force']}' — {correct}",
                correct)

# ═══════════════════════════════════════════════════════════════════════════════
# G. CHRONOLOGICAL PARADOXES within a single row
# ═══════════════════════════════════════════════════════════════════════════════
for r in rows:
    ds = parse_date(r.get("date_start",""))
    de = parse_date(r.get("date_end",""))
    if ds and de and de < ds:
        add("G. Date Paradox", "CRITICAL", [r["id"]], "date_end",
            f"date_end {r['date_end']} is BEFORE date_start {r['date_start']}",
            "Fix dates — end cannot precede start.")

# Events dated before the war
for r in rows:
    ds = parse_date(r.get("date_start",""))
    WAR_START = datetime(1899,10,11)
    if ds and ds < WAR_START:
        et = r.get("event_type","")
        if et not in ("deployment","garrison","disembark"):
            add("G. Pre-war Event", "HIGH", [r["id"]], "date_start",
                f"Event dated {r['date_start']} — before war started (11 Oct 1899). event_type='{et}'",
                "War started 11 Oct 1899 with Boer ultimatum expiry")

# ═══════════════════════════════════════════════════════════════════════════════
# H. THEATER MISMATCH — sudden appearance in wrong theater without movement row
# Rough bounding boxes: EC = lat -34 to -29, lon 22 to 30
#                       Natal = lat -29 to -27, lon 29 to 32
#                       OFS/Transvaal = lat -22 to -29, lon 24 to 32
# Flag if a unit skips theaters (e.g. EC → Natal → EC) without movement rows
# ═══════════════════════════════════════════════════════════════════════════════
def theater(lon, lat):
    if lat < -29 and lon < 30: return "Cape/EC"
    if lat > -29 and lon > 29: return "Natal"
    if lat < -26 and lon > 25: return "OFS"
    return "Transvaal/North"

for force, timeline in force_timeline.items():
    timeline.sort(key=lambda e: e["date"])
    for i in range(len(timeline)-1):
        a, b = timeline[i], timeline[i+1]
        th_a = theater(a["coords"][0], a["coords"][1])
        th_b = theater(b["coords"][0], b["coords"][1])
        if th_a != th_b:
            d1 = a.get("date_end") or a["date"]
            d2 = b["date"]
            days = (d2 - d1).days
            dist = haversine(a["coords"], b["coords"])
            if days < 14 and dist > 200:  # rapid theater change with no movement row
                add("H. Theater Jump", "MEDIUM",
                    [a["id"], b["id"]], "place/theater",
                    f"'{force}': jumps from {th_a} ({a['place']}) to {th_b} ({b['place']}) "
                    f"in {days}d / {dist:.0f}km — no movement row bridging this",
                    "Either a rail movement row is missing, or one of the place coordinates is wrong")

# ═══════════════════════════════════════════════════════════════════════════════
# I. DESCRIPTION CONTRADICTION — outcome words vs side/event_type
# ═══════════════════════════════════════════════════════════════════════════════
for r in rows:
    desc = r.get("description","")
    side = r.get("side","")
    et   = r.get("event_type","")

    # British row says "British surrendered / captured by Boers"
    if side=="British" and et not in ("surrender","defeat","retreat") and \
       re.search(r"british.*surrender|british.*captured|british.*prisoner|british.*defeat", desc, re.I):
        add("I. Outcome Contradiction", "MEDIUM", [r["id"]], "event_type",
            f"Description mentions British surrender/defeat but event_type='{et}': \"{desc[:120]}\"",
            "Should this be event_type='defeat' or 'surrender'?")

    # Boer row says "Boers surrendered / captured"
    if side=="Boer" and et not in ("surrender","defeat","retreat","capture") and \
       re.search(r"boer.*surrender|boer.*captured|boer.*defeat|boer.*routed", desc, re.I):
        add("I. Outcome Contradiction", "MEDIUM", [r["id"]], "event_type",
            f"Description mentions Boer defeat/surrender but event_type='{et}': \"{desc[:120]}\"",
            "Should this be event_type='defeat' or 'surrender'?")

# ═══════════════════════════════════════════════════════════════════════════════
# J. KNOWN COORDINATE CHECKS — famous battles must be close to true location
# (catches OVERRIDES or TOWNS entries that are clearly wrong)
# ═══════════════════════════════════════════════════════════════════════════════
COORD_CHECKS = [
    # (place_name_in_data, true_lat, true_lon, max_km, note)
    ("Spion Kop",      -28.57,  29.53, 30, "Spion Kop is in Natal, ~35km W of Ladysmith"),
    ("Colenso",        -28.73,  29.83, 20, "Colenso is on Tugela River, Natal"),
    ("Stormberg",      -31.26,  26.57, 25, "Stormberg is in EC, near Molteno"),
    ("Magersfontein",  -29.03,  24.88, 20, "Magersfontein is S of Kimberley"),
    ("Paardeberg",     -29.05,  25.00, 20, "Paardeberg is on Modder River, OFS"),
    ("Sannaspos",      -29.10,  26.27, 25, "Sannaspos/Koorn Spruit is E of Bloemfontein"),
    ("Elands River Poort", -32.32, 24.97, 30, "Elands River Poort is in Bamboesberg, EC"),
    ("Kimberley",      -28.73,  24.77, 20, "Kimberley, Northern Cape"),
    ("Mafeking",       -25.85,  25.65, 20, "Mafeking (Mahikeng), NW Province"),
    ("Ladysmith",      -28.55,  29.77, 15, "Ladysmith, KZN"),
    ("Bloemfontein",   -29.12,  26.21, 20, "Bloemfontein, OFS"),
    ("Pretoria",       -25.75,  28.23, 20, "Pretoria, Transvaal"),
    ("Nooitgedacht",   -25.73,  27.49, 25, "Nooitgedacht is in Magaliesberg, Transvaal"),
    ("Tweebosch",      -26.81,  23.63, 30, "Tweebosch is NW Transvaal, near Vryburg"),
    ("Diamond Hill",   -25.79,  28.46, 20, "Diamond Hill (Donkerhoek) is E of Pretoria"),
    ("Bergendal",      -25.78,  30.24, 20, "Bergendal is near Belfast, E Transvaal"),
]

for place_name, true_lat, true_lon, max_km, note in COORD_CHECKS:
    for fid, feat in feat_by_id.items():
        p = feat["properties"]
        if p.get("place","").lower() == place_name.lower():
            coords = feat["geometry"]["coordinates"]
            dist = haversine(coords, [true_lon, true_lat])
            if dist > max_km:
                add("J. Wrong Coordinates", "HIGH", [fid], "coordinates",
                    f"'{place_name}' marker is {dist:.0f}km from true location "
                    f"(marker: [{coords[1]:.3f},{coords[0]:.3f}], "
                    f"should be ~[{true_lat:.3f},{true_lon:.3f}])",
                    note)

# ═══════════════════════════════════════════════════════════════════════════════
# K. DUPLICATE EVENTS — same place + date + side combination
# ═══════════════════════════════════════════════════════════════════════════════
event_sig = defaultdict(list)
for r in rows:
    place = r.get("action_place","").strip()
    date  = r.get("date_start","").strip()
    side  = r.get("side","").strip()
    et    = r.get("event_type","").strip()
    if place and date and side:
        key = (place, date, side, et)
        event_sig[key].append(r["id"])

for key, ids in event_sig.items():
    if len(ids) > 1:
        place, date, side, et = key
        add("K. Possible Duplicate", "MEDIUM", ids,
            "place/date/side",
            f"Multiple rows: {side} '{et}' at {place} on {date} (IDs: {', '.join(ids)})",
            "Check if these are truly separate events or the same event entered twice")

# ═══════════════════════════════════════════════════════════════════════════════
# L. EVENT AFTER PEACE — any non-diplomatic events after 31 May 1902
# ═══════════════════════════════════════════════════════════════════════════════
PEACE = datetime(1902,5,31)
for r in rows:
    ds = parse_date(r.get("date_start",""))
    et = r.get("event_type","")
    if ds and ds > PEACE and et not in ("redeployment","command","exile","occupation"):
        add("L. Post-peace Event", "HIGH", [r["id"]], "date_start",
            f"Active military event ('{et}') dated {r['date_start']} — after Peace of Vereeniging (31 May 1902)",
            "War ended 31 May 1902. Events after this date should be redeployment/occupation/exile only.")

# ═══════════════════════════════════════════════════════════════════════════════
# M. SPECIFIC KNOWN ERRORS — hardcoded checks for specific historical facts
# ═══════════════════════════════════════════════════════════════════════════════
KNOWN_FACTS = [
    # (row_id, field, wrong_value_pattern, note)
    # Paardeberg: Cronjé surrendered 27 Feb 1900 — Majuba Day
    # Ladysmith relief: 28 Feb 1900 (same day as Paardeberg surrender, different front)
    # Mafeking: 217 days, 12 Oct 1899 – 17 May 1900
    # Kimberley siege: 14 Oct 1899 – 15 Feb 1900 = 124 days
    # Wepener siege: 9–25 April 1900 (16 days)
    # Spion Kop: British had 243 killed — bloodiest single day of the war
    # Nooitgedacht: 13 Dec 1900, Clements ambushed (NOT Clements surrendered — he escaped)
    # Peace signed: Vereeniging, 31 May 1902 (not 30 May or 1 Jun)
]

# Check specific known dates
DATE_FACTS = {
    # keywords in force/desc → (field, wrong_pattern, correct_date, note)
    "paardeberg.*surrender|cronje.*surrender|cronjé.*surrender": ("date_start", r"1900-02-(?!27)", "1900-02-27", "Cronjé surrendered specifically on 27 Feb 1900 (Majuba Day)"),
    "mafeking.*reliev|reliev.*mafeking": ("date_start", r"1900-05-(?!17)", "1900-05-17", "Mafeking relieved 17 May 1900"),
    "kimberley.*reliev|reliev.*kimberley": ("date_start", r"1900-02-(?!15)", "1900-02-15", "Kimberley relieved 15 Feb 1900 by French's cavalry"),
    "ladysmith.*reliev|reliev.*ladysmith": ("date_start", r"1900-02-(?!28)", "1900-02-28", "Ladysmith relieved 28 Feb 1900"),
    "spion kop|spionkop": ("date_start", r"1900-01-(?!24)", "1900-01-24", "Battle of Spion Kop was 24 Jan 1900"),
    "stormberg": ("date_start", r"1899-12-(?!10)", "1899-12-10", "Battle of Stormberg was 10 Dec 1899"),
    "magersfontein": ("date_start", r"1899-12-(?!11)", "1899-12-11", "Battle of Magersfontein was 11 Dec 1899"),
    "colenso": ("date_start", r"1899-12-(?!15)", "1899-12-15", "Battle of Colenso was 15 Dec 1899"),
    "vereeniging": ("date_start", r"1902-05-(?!31)", "1902-05-31", "Peace of Vereeniging signed 31 May 1902"),
    "nooitgedacht": ("date_start", r"1900-12-(?!13)", "1900-12-13", "Nooitgedacht was 13 Dec 1900"),
    "bergendal": ("date_start", r"1900-08-(?!27)", "1900-08-27", "Bergendal (last pitched battle) was 27 Aug 1900"),
}

for r in rows:
    combined = (r.get("force","") + " " + r.get("description","")).lower()
    for pattern, (field, wrong_pat, correct, note) in DATE_FACTS.items():
        if re.search(pattern, combined):
            val = r.get(field,"")
            if val and re.search(wrong_pat, val):
                add("M. Known Wrong Date", "CRITICAL", [r["id"]], field,
                    f"'{r['force']}': {field}='{val}' but should be {correct}",
                    note)

# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY + EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
from collections import Counter
by_check = defaultdict(list)
by_sev   = Counter()
for iss in issues:
    by_check[iss["check"]].append(iss)
    by_sev[iss["severity"]] += 1

print(f"LOGICAL AUDIT COMPLETE")
print(f"Total issues: {len(issues)}")
for s in ["CRITICAL","HIGH","MEDIUM","LOW"]:
    print(f"  {s}: {by_sev[s]}")
print()
for check, lst in sorted(by_check.items()):
    print(f"  {check}: {len(lst)}")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
SEV_COL = {"CRITICAL":"CC0000","HIGH":"FF6600","MEDIUM":"FFAA00","LOW":"AAAAAA"}

ws = wb.active; ws.title = "Logical Issues"
HDR = ["Severity","Check","Row IDs","Force","Field","Description","Historical Note"]
ws.append(HDR)
hf = PatternFill("solid",fgColor="0d1b2a"); hfn=Font(bold=True,color="FFFFFF",size=10)
for c in range(1,8): ws.cell(1,c).fill=hf; ws.cell(1,c).font=hfn

ORDER = ["CRITICAL","HIGH","MEDIUM","LOW"]
for iss in sorted(issues, key=lambda x:(ORDER.index(x["severity"]),x["check"])):
    ws.append([iss["severity"],iss["check"],iss["ids"],iss["force"],iss["field"],iss["description"],iss["historical_note"]])
    r = ws.max_row
    col = SEV_COL.get(iss["severity"],"CCCCCC")
    ws.cell(r,1).fill=PatternFill("solid",fgColor=col)
    ws.cell(r,1).font=Font(bold=True,color="FFFFFF" if iss["severity"] in ("CRITICAL","HIGH") else "333333")
    ws.cell(r,6).alignment=Alignment(wrap_text=True)
    ws.cell(r,7).alignment=Alignment(wrap_text=True)
    ws.cell(r,7).font=Font(italic=True,color="1a4ab9")

WIDTHS = [10,22,12,35,14,70,55]
for i,w in enumerate(WIDTHS,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions

# ── Per-check sheets ──────────────────────────────────────────────────────────
for check, lst in sorted(by_check.items()):
    safe = re.sub(r"[^\w ]","",check)[:31]
    ws2 = wb.create_sheet(safe)
    ws2.append(["Severity","Row IDs","Force","Field","Description","Historical Note"])
    hf2=PatternFill("solid",fgColor="1a2a3a"); hfn2=Font(bold=True,color="FFFFFF",size=10)
    for c in range(1,7): ws2.cell(1,c).fill=hf2; ws2.cell(1,c).font=hfn2
    for iss in sorted(lst, key=lambda x:ORDER.index(x["severity"])):
        ws2.append([iss["severity"],iss["ids"],iss["force"],iss["field"],iss["description"],iss["historical_note"]])
        r=ws2.max_row
        ws2.cell(r,1).fill=PatternFill("solid",fgColor=SEV_COL.get(iss["severity"],"CCCCCC"))
        ws2.cell(r,1).font=Font(bold=True,color="FFFFFF" if iss["severity"] in ("CRITICAL","HIGH") else "333333")
        ws2.cell(r,5).alignment=Alignment(wrap_text=True)
        ws2.cell(r,6).alignment=Alignment(wrap_text=True)
    for c,w in enumerate([10,12,35,14,70,55],1): ws2.column_dimensions[get_column_letter(c)].width=w
    ws2.freeze_panes="A2"

out = HERE/"tools"/"boer_war_logical_audit.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
